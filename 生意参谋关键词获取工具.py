import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox
import threading
import os
import time
import re
import shutil
import traceback
import subprocess
import win32com.client  # 需要安装pywin32库：pip install pywin32
try:
    import pythoncom
except ImportError:  # 环境缺少pywin32组件时兜底
    pythoncom = None
import openpyxl
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from typing import Optional
import math
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
import warnings

# 抑制libpng警告
warnings.filterwarnings("ignore", category=UserWarning, message="libpng warning: iCCP: known incorrect sRGB profile")


@dataclass
class BrowserSessionState:
    session_id: int
    port: int
    profile_path: str
    browser: Optional[webdriver.Chrome] = None
    categories: dict = field(default_factory=lambda: {1: [], 2: [], 3: []})
    selected_categories: dict = field(default_factory=lambda: {1: None, 2: None, 3: None})
    collected_data: list = field(default_factory=list)
    current_level: int = 1
    current_level1_index: int | None = None
    excel_filepath: str = ""
    excel_app: any = None
    current_excel_root: str = ""
    interface_opened: bool = False
    processing: bool = False
    paused: bool = False
    pause_event: threading.Event = field(default_factory=threading.Event)
    stop_event: threading.Event = field(default_factory=threading.Event)
    output_dir: str = ""
    excel_dirty: bool = False
    level1_total: int = 0
    level1_current: int = 0
    level2_total: int = 0
    level2_current: int = 0

    def __post_init__(self):
        self.pause_event.set()


class CategoryAutoExtractor(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("搜索词自动提取工具")
        self.configure(bg="#f0f0f0")
        self.width = 800
        self.height = 650
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - self.width) // 2
        y = (screen_height - self.height) // 2
        self.geometry(f"{self.width}x{self.height}+{x}+{y}")

        self.initialize_other_properties()
        self.create_widgets()

        # 新增：进度条相关状态变量
        self.total_level1 = 0  # 总一级类目数
        self.current_level1 = 0  # 当前处理的一级类目索引
        self.total_level2 = 0  # 当前一级类目下的总二级类目数
        self.current_level2 = 0  # 当前处理的二级类目索引

        # 新增：日志优化相关配置
        self.log_buffer = []  # 日志缓冲区
        self.log_batch_size = 5  # 批量更新阈值（达到5条日志则更新）
        self.log_max_lines = 5000  # 最大日志行数，超过则清理旧日志
        self.log_updating = False  # 防止重复触发批量更新

        # 线程事件状态控制由会话维护
        # 设置窗口关闭协议
        self.protocol('WM_DELETE_WINDOW', self.on_closing)

    def initialize_other_properties(self):
        """初始化非UI属性"""
        self.exclude_level1_serials = [4, 34, 52, 53, 54, 58, 59, 60]  # 示例：排除第3、5、8个一级类目
        self.log_ui(f'已配置排除一级类目序号：{self.exclude_level1_serials}')  # 日志提示
        # 多浏览器调试配置
        self.base_debug_port = 9000
        self.port_step = 100
        self.debug_profile_root = r"C:/temp/chrome_debug_profiles"
        os.makedirs(self.debug_profile_root, exist_ok=True)

        # 数据收集相关配置
        self.min_popularity_threshold = 150  # 筛选搜索人气大于等于此值的数据
        self.max_pages = 6  # 最多处理的页数
        self.stop_on_low_value = True  # 遇到小于阈值的值时停止
        self.auto_total_categories = 61
        self.auto_block_size = 10

        # 全局状态
        self.driver_path = ''
        self.cached_driver_path = None  # 缓存已验证的驱动路径
        self.output_root = os.path.join(r"D:\Desktop\python\gjc", 'dist', '汇总')
        os.makedirs(self.output_root, exist_ok=True)
        self.lock = threading.Lock()
        self.closing = False
        self.session_states = {}
        self.active_session_id = -1
        self._session_local = threading.local()
        self._default_session_state = self._create_session_state(session_id=-1, port=self.base_debug_port)
        self.session_states[-1] = self._default_session_state
        self._session_local.state = self._default_session_state
        self.cookie_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cookie.txt")

        # 初始化完成后显示提示
        self.log_ui('程序初始化完成，请先配置Chrome驱动路径')

    def _create_session_state(self, session_id: int, port: int) -> BrowserSessionState:
        """按端口创建独立会话状态"""
        profile_path = os.path.join(self.debug_profile_root, f"profile_{port}")
        os.makedirs(profile_path, exist_ok=True)

        output_dir = self.output_root
        os.makedirs(output_dir, exist_ok=True)

        state = BrowserSessionState(
            session_id=session_id,
            port=port,
            profile_path=profile_path,
            output_dir=output_dir,
        )
        return state

    def _ensure_session_state(self, session_id: int) -> BrowserSessionState:
        if session_id not in self.session_states:
            port = self.base_debug_port + session_id * self.port_step
            self.session_states[session_id] = self._create_session_state(session_id, port)
        return self.session_states[session_id]

    def set_session_status(self, session_id, status):
        panel = getattr(self, "session_panels", {}).get(session_id)
        if panel:
            panel["status_var"].set(status)


    def _bind_session_state(self, state: BrowserSessionState):
        """在当前线程绑定指定会话"""
        self._session_local.state = state

    def _get_active_state(self) -> BrowserSessionState:
        state = getattr(self._session_local, "state", None)
        if state is None:
            if self.active_session_id in self.session_states:
                state = self.session_states[self.active_session_id]
            else:
                state = self._default_session_state
            self._session_local.state = state
        return state

    def _set_active_session(self, session_id: int):
        """在UI线程切换当前活跃会话"""
        if session_id in self.session_states:
            self.active_session_id = session_id
            self._bind_session_state(self.session_states[session_id])

    def _iter_real_sessions(self):
        """遍历所有真实浏览器会话（排除默认占位）"""
        for session_id, state in self.session_states.items():
            if session_id >= 0:
                yield state

    def _run_in_session_thread(self, state: BrowserSessionState, target, *args, **kwargs):
        """在独立线程中运行会话相关任务"""
        def runner():
            com_initialized = False
            if pythoncom:
                try:
                    pythoncom.CoInitialize()
                    com_initialized = True
                except Exception as exc:  # pylint: disable=broad-except
                    self.log_console(f"COM初始化失败: {exc}")
            self._bind_session_state(state)
            try:
                target(*args, **kwargs)
            except Exception as exc:  # pylint: disable=broad-except
                self.log_ui(f"会话{state.session_id + 1} 执行任务失败: {exc}")
                self.log_console(traceback.format_exc())
            finally:
                if com_initialized:
                    try:
                        pythoncom.CoUninitialize()
                    except Exception as exc:  # pylint: disable=broad-except
                        self.log_console(f"COM释放失败: {exc}")
        threading.Thread(target=runner, daemon=True).start()

    def _get_ready_sessions(self, require_interface=False):
        sessions = [state for state in self._iter_real_sessions() if state.browser]
        if require_interface:
            sessions = [state for state in sessions if state.interface_opened]
        return sessions

    def _split_range(self, start: int, end: int, parts: int):
        total = end - start + 1
        if parts <= 0:
            return []
        chunk = max(1, math.ceil(total / parts))
        ranges = []
        current = start
        while current <= end and len(ranges) < parts:
            chunk_end = min(end, current + chunk - 1)
            ranges.append((current, chunk_end))
            current = chunk_end + 1
        return ranges

    def _split_indices_evenly(self, indices, parts):
        """将索引列表平均切分为parts份，保持原有顺序"""
        result = []
        total = len(indices)
        if parts <= 0:
            return result
        base = total // parts
        extra = total % parts
        start = 0
        for i in range(parts):
            length = base + (1 if i < extra else 0)
            if length <= 0:
                result.append([])
                continue
            result.append(indices[start:start + length])
            start += length
        return result

    def _start_processing_task(self, state: BrowserSessionState, target, *args):
        state.stop_event.clear()
        state.pause_event.set()
        state.processing = True
        state.paused = False

        def runner():
            try:
                target(*args)
            finally:
                state.processing = False
                state.paused = False
                state.stop_event.clear()
                self.after(0, lambda sid=state.session_id: self._on_session_processing_finished(sid))

        panel = getattr(self, "session_panels", {}).get(state.session_id)
        if panel:
            panel["pause_btn"].config(state=tk.NORMAL, text="暂停")
            panel["stop_btn"].config(state=tk.NORMAL)
        self.set_session_status(state.session_id, "运行中")
        self._run_in_session_thread(state, runner)

    def _on_session_processing_finished(self, session_id):
        panel = getattr(self, "session_panels", {}).get(session_id)
        state = self.session_states.get(session_id)
        if panel and (not state or not state.processing):
            panel["pause_btn"].config(state=tk.DISABLED, text="暂停")
            panel["stop_btn"].config(state=tk.DISABLED)
        if state and not state.processing:
            self.set_session_status(session_id, "准备完成")

    def log_session(self, message: str):
        """带会话前缀的日志输出"""
        state = self._get_active_state()
        prefix = ""
        if state.session_id >= 0:
            prefix = f"[窗口{state.session_id + 1}] "
        self.log_ui(prefix + message)

    def _reset_collected_data(self):
        self.collected_data = []

    def _append_collected_data(self, page_data):
        if page_data:
            self.collected_data.extend(page_data)
            self.save_page_data_to_excel(page_data, commit=False)

    def _flush_collected_data(self):
        state = self._get_active_state()
        had_data = bool(self.collected_data) or getattr(state, "excel_dirty", False)
        self.collected_data = []
        if getattr(state, "excel_dirty", False):
            self.save_page_data_to_excel([], commit=True)
        return had_data

    # ===== 会话属性访问器 =====
    @property
    def browser(self):
        return self._get_active_state().browser

    @browser.setter
    def browser(self, value):
        self._get_active_state().browser = value

    @property
    def categories(self):
        return self._get_active_state().categories

    @categories.setter
    def categories(self, value):
        self._get_active_state().categories = value

    @property
    def selected_categories(self):
        return self._get_active_state().selected_categories

    @selected_categories.setter
    def selected_categories(self, value):
        self._get_active_state().selected_categories = value

    @property
    def collected_data(self):
        return self._get_active_state().collected_data

    @collected_data.setter
    def collected_data(self, value):
        self._get_active_state().collected_data = value

    @property
    def current_level(self):
        return self._get_active_state().current_level

    @current_level.setter
    def current_level(self, value):
        self._get_active_state().current_level = value

    @property
    def current_level1_index(self):
        return self._get_active_state().current_level1_index

    @current_level1_index.setter
    def current_level1_index(self, value):
        self._get_active_state().current_level1_index = value

    @property
    def excel_filepath(self):
        return self._get_active_state().excel_filepath

    @excel_filepath.setter
    def excel_filepath(self, value):
        self._get_active_state().excel_filepath = value

    @property
    def excel_app(self):
        return self._get_active_state().excel_app

    @excel_app.setter
    def excel_app(self, value):
        self._get_active_state().excel_app = value

    @property
    def current_excel_root(self):
        return self._get_active_state().current_excel_root

    @current_excel_root.setter
    def current_excel_root(self, value):
        self._get_active_state().current_excel_root = value

    @property
    def interface_opened(self):
        return self._get_active_state().interface_opened

    @interface_opened.setter
    def interface_opened(self, value):
        self._get_active_state().interface_opened = value

    @property
    def processing(self):
        return self._get_active_state().processing

    @processing.setter
    def processing(self, value):
        self._get_active_state().processing = value

    @property
    def paused(self):
        return self._get_active_state().paused

    @paused.setter
    def paused(self, value):
        self._get_active_state().paused = value

    @property
    def pause_event(self):
        return self._get_active_state().pause_event

    @pause_event.setter
    def pause_event(self, value):
        self._get_active_state().pause_event = value

    @property
    def stop_event(self):
        return self._get_active_state().stop_event

    @stop_event.setter
    def stop_event(self, value):
        self._get_active_state().stop_event = value

    @property
    def output_dir(self):
        return self._get_active_state().output_dir

    @output_dir.setter
    def output_dir(self, value):
        self._get_active_state().output_dir = value
    def create_widgets(self):
        """创建UI组件"""
        try:
            path_frame = ttk.LabelFrame(self, text="浏览器配置", padding=(10, 5))
            path_frame.pack(fill=tk.X, padx=10, pady=(10, 5))

            ttk.Label(path_frame, text="Chrome驱动路径:").pack(side=tk.LEFT, padx=(0, 5))

            self.driver_path_var = tk.StringVar(
                value=r'D:\Desktop\python\gjc\chromedriver-win64\chromedriver.exe')
            ttk.Entry(path_frame, textvariable=self.driver_path_var, width=50).pack(
                side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10)
            )

            ttk.Button(path_frame, text="浏览", command=self.browse_driver).pack(side=tk.RIGHT)

            window_frame = ttk.Frame(path_frame)
            window_frame.pack(fill=tk.X, pady=(5, 0))
            ttk.Label(window_frame, text="窗口数量:").pack(side=tk.LEFT, padx=(0, 5))
            self.window_count_var = tk.IntVar(value=1)
            self.window_count_spin = tk.Spinbox(
                window_frame,
                from_=1,
                to=20,
                textvariable=self.window_count_var,
                width=5,
                command=self.on_window_count_change
            )
            self.window_count_spin.pack(side=tk.LEFT)
            ttk.Label(window_frame, text="(端口 9000 起，每 +100)").pack(side=tk.LEFT, padx=(5, 0))
            self.window_count_var.trace_add("write", lambda *args: self.on_window_count_change())

            self.group_control_frame = ttk.LabelFrame(self, text="群控操作", padding=(10, 5))
            self._build_group_controls()

            self.sessions_outer = ttk.Frame(self)
            self.sessions_outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))

            self.sessions_canvas = tk.Canvas(self.sessions_outer, highlightthickness=0)
            self.sessions_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar = ttk.Scrollbar(self.sessions_outer, orient=tk.VERTICAL, command=self.sessions_canvas.yview)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            self.sessions_canvas.configure(yscrollcommand=scrollbar.set)

            self.sessions_inner = ttk.Frame(self.sessions_canvas)
            self.sessions_window = self.sessions_canvas.create_window((0, 0), window=self.sessions_inner, anchor="nw")

            def _sync_scrollregion(event):
                self.sessions_canvas.configure(scrollregion=self.sessions_canvas.bbox("all"))
                self.sessions_canvas.itemconfig(self.sessions_window, width=event.width)

            self.sessions_inner.bind("<Configure>", _sync_scrollregion)
            self.sessions_canvas.bind("<Configure>", _sync_scrollregion)

            def _on_mousewheel(event):
                if event.delta:
                    self.sessions_canvas.yview_scroll(int(-event.delta / 120), "units")

            self.sessions_canvas.bind("<MouseWheel>", _on_mousewheel)
            self.sessions_inner.bind("<MouseWheel>", _on_mousewheel)

            self.session_panels = {}
            self.build_session_cards()
            self.update_group_controls_visibility()

            self.log_ui("UI组件创建成功")

        except Exception as e:
            error_msg = f"UI组件创建失败: {str(e)}\n{traceback.format_exc()}"
            print(error_msg)
            error_window = tk.Toplevel(self)
            error_window.title("UI初始化错误")
            error_window.geometry("400x300")
            tk.Label(error_window, text="UI加载失败，请检查环境配置：").pack(padx=10, pady=10)
            error_text = scrolledtext.ScrolledText(error_window, wrap=tk.WORD)
            error_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            error_text.insert(tk.END, error_msg)

    def browse_driver(self):
        """浏览选择Chrome驱动路径"""
        path = filedialog.askopenfilename(
            title="选择Chrome驱动",
            filetypes=[("可执行文件", "*.exe"), ("所有文件", "*.*")]
        )
        if path:
            self.driver_path_var.set(path)
            self.log_ui(f"已选择Chrome驱动路径: {path}")

    def on_window_count_change(self):
        """窗口数量调整时刷新会话UI"""
        try:
            count = int(self.window_count_var.get())
        except Exception:
            count = 1
        count = max(1, min(20, count))
        self.build_session_cards()
        self.update_group_controls_visibility()

    def _build_group_controls(self):
        row = ttk.Frame(self.group_control_frame)
        row.pack(fill=tk.X)
        ttk.Button(row, text="打开浏览器", command=self.group_open_browser).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(row, text="打开界面", command=self.group_open_interface).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(row, text="Cookie登录", command=self.group_cookie_login).pack(side=tk.LEFT, padx=(0, 5))
        self.group_pause_btn = ttk.Button(row, text="暂停", width=10, command=self.group_toggle_pause)
        self.group_pause_btn.pack(side=tk.LEFT, padx=(5, 5))
        self.group_stop_btn = ttk.Button(row, text="结束", width=10, command=self.group_stop_all)
        self.group_stop_btn.pack(side=tk.LEFT, padx=(0, 5))

        # 群控类目分配输入
        self.group_entry = ttk.Entry(row, width=10)
        self.group_entry.pack(side=tk.LEFT, padx=(10, 5))
        self.group_entry.bind("<Return>", self.group_process_input)
        ttk.Button(row, text="确认", command=self.group_process_input).pack(side=tk.LEFT, padx=(0, 5))

        # 群控 Cookie 输入（回车即保存并登录，带占位提示）
        self.group_cookie_entry = ttk.Entry(row, width=30)
        self.group_cookie_placeholder = "cookie"
        self.group_cookie_placeholder_color = "#999999"
        self.group_cookie_normal_color = self.group_cookie_entry.cget("foreground")
        self.group_cookie_entry.pack(side=tk.LEFT, padx=(10, 5))
        self.group_cookie_entry.bind("<Return>", self.group_cookie_input_submit)
        self.group_cookie_entry.bind("<FocusIn>", self._handle_cookie_focus_in)
        self.group_cookie_entry.bind("<FocusOut>", self._handle_cookie_focus_out)
        self._set_cookie_placeholder()

    def update_group_controls_visibility(self):
        if not hasattr(self, "group_control_frame"):
            return
        try:
            count = int(self.window_count_var.get())
        except Exception:
            count = 1
        self.group_control_frame.pack_forget()
        if count > 1 and hasattr(self, "sessions_outer"):
            self.group_control_frame.pack(fill=tk.X, padx=10, pady=(0, 5), before=self.sessions_outer)
        if hasattr(self, "sessions_outer"):
            self.sessions_outer.pack_forget()
            self.sessions_outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))
            self.after_idle(self._refresh_sessions_layout)

    def _refresh_sessions_layout(self):
        if not hasattr(self, "sessions_canvas") or not hasattr(self, "sessions_window"):
            return
        self.sessions_canvas.update_idletasks()
        width = self.sessions_canvas.winfo_width()
        if width > 0:
            self.sessions_canvas.itemconfig(self.sessions_window, width=width)
        self.sessions_canvas.configure(scrollregion=self.sessions_canvas.bbox("all"))

    def _collect_group_sessions(self, require_browser=False, require_interface=False):
        count = max(1, int(self.window_count_var.get() or 1))
        sessions = []
        for sid in range(count):
            state = self._ensure_session_state(sid)
            if require_browser and not state.browser:
                continue
            if require_interface and not state.interface_opened:
                continue
            sessions.append(state)
        return sessions

    def group_open_browser(self):
        for state in self._collect_group_sessions():
            self.open_browser(state.session_id)

    def group_open_interface(self):
        sessions = self._collect_group_sessions(require_browser=True)
        if not sessions:
            self.log_ui("请先打开至少一个浏览器")
            return
        for state in sessions:
            self.open_interface(state.session_id)

    def group_cookie_login(self):
        sessions = self._collect_group_sessions(require_browser=True)
        if not sessions:
            self.log_ui("无可用窗口执行Cookie登录")
            return
        for state in sessions:
            self.cookie_login(state.session_id)

    def group_toggle_pause(self):
        sessions = [state for state in self._iter_real_sessions() if state.processing]
        if not sessions:
            self.log_ui("当前没有运行中的会话")
            return
        should_pause = not all(state.paused for state in sessions)
        if should_pause:
            for state in sessions:
                if not state.paused:
                    self.toggle_pause(state.session_id)
            self.group_pause_btn.config(text="继续")
            self.log_ui("已群控暂停所有运行中的会话")
        else:
            for state in sessions:
                if state.paused:
                    self.toggle_pause(state.session_id)
            self.group_pause_btn.config(text="暂停")
            self.log_ui("已恢复所有暂停的会话")

    def group_stop_all(self):
        sessions = [state for state in self._iter_real_sessions() if state.processing]
        if not sessions:
            self.log_ui("没有需要结束的任务")
            return
        for state in sessions:
            self.stop_processing(state.session_id)
        self.group_pause_btn.config(text="暂停")
        self.log_ui("已发出结束指令")

    def _sync_categories_to_sessions(self, reference_state, sessions):
        for state in sessions:
            if state is reference_state:
                continue
            for level in (1, 2, 3):
                state.categories[level] = [dict(item) for item in reference_state.categories[level]]

    def group_process_input(self, event=None):
        if not hasattr(self, "group_entry"):
            return
        input_text = self.group_entry.get().strip()
        self.group_entry.delete(0, tk.END)
        if not input_text:
            self.log_ui("请输入群控指令")
            return

        sessions = self._collect_group_sessions(require_browser=True, require_interface=True)
        if not sessions:
            self.log_ui("请先为所有窗口打开浏览器并加载界面")
            return

        sessions.sort(key=lambda s: s.session_id)
        reference_state = sessions[0]
        if not reference_state.categories[1]:
            self.log_ui("请先在第一个窗口加载一级类目")
            return
        self._sync_categories_to_sessions(reference_state, sessions)

        if input_text == "0":
            indices = [cat['index'] for cat in reference_state.categories[1]
                       if cat['index'] not in self.exclude_level1_serials]
        elif '-' in input_text:
            try:
                start_idx, end_idx = map(int, input_text.split('-'))
            except ValueError:
                self.log_ui("范围格式错误，请使用5-8形式")
                return
            if start_idx < 1 or end_idx < start_idx:
                self.log_ui("范围输入无效")
                return
            indices = [cat['index'] for cat in reference_state.categories[1]
                       if start_idx <= cat['index'] <= end_idx
                       and cat['index'] not in self.exclude_level1_serials]
        else:
            self.log_ui("群控仅支持输入0或范围(如5-8)")
            return

        if not indices:
            self.log_ui("没有符合条件的一级类目可分配")
            return

        chunks = self._split_indices_evenly(indices, len(sessions))
        for state, chunk in zip(sessions, chunks):
            if not chunk:
                self.log_ui(f"窗口{state.session_id + 1} 未分配到类目")
                continue
            self.log_ui(f"窗口{state.session_id + 1} 负责类目: {chunk}")
            self._start_processing_task(state, self.process_assigned_categories, chunk)

    def group_cookie_input_submit(self, event=None):
        if not hasattr(self, "group_cookie_entry"):
            return
        raw = self.group_cookie_entry.get().strip()
        self.group_cookie_entry.delete(0, tk.END)
        if raw == self.group_cookie_placeholder:
            raw = ""
        if not raw:
            self.log_ui("请输入Cookie字符串")
            self._set_cookie_placeholder()
            return
        try:
            os.makedirs(os.path.dirname(self.cookie_path), exist_ok=True)
            with open(self.cookie_path, "w", encoding="utf-8") as f:
                f.write(raw)
            self.log_ui(f"Cookie已写入文件: {self.cookie_path}")
            # 写入后直接群控登录
            self.group_cookie_login()
        except Exception as exc:  # pylint: disable=broad-except
            self.log_ui(f"保存或登录Cookie时出错: {exc}")

    def _set_cookie_placeholder(self):
        self.group_cookie_entry.delete(0, tk.END)
        self.group_cookie_entry.insert(0, self.group_cookie_placeholder)
        self.group_cookie_entry.config(foreground=self.group_cookie_placeholder_color)
        self._cookie_placeholder_active = True

    def _handle_cookie_focus_in(self, event):
        if getattr(self, "_cookie_placeholder_active", False):
            event.widget.delete(0, tk.END)
            event.widget.config(foreground=self.group_cookie_normal_color)
            self._cookie_placeholder_active = False

    def _handle_cookie_focus_out(self, event):
        if not event.widget.get().strip():
            self._set_cookie_placeholder()

    def build_session_cards(self):
        desired = max(1, int(self.window_count_var.get() or 1))
        layout_mode = "multi" if desired > 1 else "single"
        panels = getattr(self, "session_panels", {})
        for sid in list(panels.keys()):
            if sid >= desired:
                panel = panels.pop(sid)
                panel["frame"].destroy()
        for sid, panel in list(panels.items()):
            if panel.get("layout_mode") != layout_mode:
                panel["frame"].destroy()
                panels.pop(sid)
        for sid in range(desired):
            if sid not in panels:
                panels[sid] = self._create_session_panel(sid, layout_mode=layout_mode)
        self.session_panels = panels
        self.sessions_inner.update_idletasks()
        self._layout_sessions_grid(desired)

    def _layout_sessions_grid(self, count):
        for c in range(2):
            self.sessions_inner.grid_columnconfigure(c, weight=0)
        if count <= 1:
            frame = self.session_panels.get(0, {}).get("frame")
            if frame:
                frame.grid_configure(row=0, column=0, sticky="nsew")
            self.sessions_inner.grid_columnconfigure(0, weight=1)
        else:
            for sid, panel in self.session_panels.items():
                row, col = divmod(sid, 2)
                panel["frame"].grid_configure(row=row, column=col, sticky="nsew", padx=5, pady=5)
                self.sessions_inner.grid_columnconfigure(col, weight=1)

    def _create_session_panel(self, session_id, layout_mode="single"):
        self._ensure_session_state(session_id)
        frame = ttk.Frame(self.sessions_inner, padding=10, relief=tk.GROOVE)
        row, col = divmod(session_id, 2)
        frame.grid(row=row, column=col, sticky="nsew", padx=5, pady=5)
        self.sessions_inner.grid_columnconfigure(col, weight=1)

        header = ttk.Frame(frame)
        header.pack(fill=tk.X)
        ttk.Label(header, text=f"会话 {session_id + 1}", font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        status_var = tk.StringVar(value="待初始化")
        ttk.Label(header, textvariable=status_var, foreground="#1a73e8").pack(side=tk.RIGHT)

        if layout_mode == "single":
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill=tk.X, pady=(5, 5))
            ttk.Button(btn_frame, text="打开浏览器",
                       command=lambda sid=session_id: self.open_browser(sid)).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(btn_frame, text="打开界面",
                       command=lambda sid=session_id: self.open_interface(sid)).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(btn_frame, text="Cookie登录",
                       command=lambda sid=session_id: self.cookie_login(sid)).pack(side=tk.LEFT, padx=(0, 5))
            pause_btn = ttk.Button(btn_frame, text="暂停",
                                   command=lambda sid=session_id: self.toggle_pause(sid), state=tk.DISABLED)
            pause_btn.pack(side=tk.LEFT, padx=(10, 5))
            stop_btn = ttk.Button(btn_frame, text="结束",
                                  command=lambda sid=session_id: self.stop_processing(sid), state=tk.DISABLED)
            stop_btn.pack(side=tk.LEFT)

            input_frame = ttk.Frame(frame)
            input_frame.pack(fill=tk.X, pady=(5, 5))
            entry = ttk.Entry(input_frame, width=18)
            entry.pack(side=tk.LEFT, padx=(0, 10))
            entry.bind("<Return>", lambda event, sid=session_id: self.process_input(sid, event))
            ttk.Button(input_frame, text="确认",
                       command=lambda sid=session_id: self.process_input(sid)).pack(side=tk.LEFT)
        else:
            grid_frame = ttk.Frame(frame)
            grid_frame.pack(fill=tk.X, pady=(5, 5))
            uniform_name = f"session_{session_id}_grid"
            for col in range(4):
                grid_frame.grid_columnconfigure(col, weight=1, uniform=uniform_name)

            ttk.Button(grid_frame, text="打开浏览器",
                       command=lambda sid=session_id: self.open_browser(sid)).grid(row=0, column=0, padx=(0, 5), sticky="ew")
            ttk.Button(grid_frame, text="打开界面",
                       command=lambda sid=session_id: self.open_interface(sid)).grid(row=0, column=1, padx=(0, 5), sticky="ew")
            ttk.Button(grid_frame, text="Cookie登录",
                       command=lambda sid=session_id: self.cookie_login(sid)).grid(row=0, column=2, padx=(0, 5), sticky="ew")
            pause_btn = ttk.Button(grid_frame, text="暂停",
                                   command=lambda sid=session_id: self.toggle_pause(sid),
                                   state=tk.DISABLED, width=10)
            pause_btn.grid(row=0, column=3, padx=(0, 5), sticky="ew")

            entry = ttk.Entry(grid_frame)
            entry.grid(row=1, column=0, columnspan=2, sticky="ew", padx=(0, 10), pady=(2, 0))
            entry.bind("<Return>", lambda event, sid=session_id: self.process_input(sid, event))
            ttk.Button(grid_frame, text="确认",
                       command=lambda sid=session_id: self.process_input(sid)).grid(row=1, column=2, sticky="ew", padx=(0, 5), pady=(2, 0))
            stop_btn = ttk.Button(grid_frame, text="结束",
                                  command=lambda sid=session_id: self.stop_processing(sid),
                                  state=tk.DISABLED, width=10)
            stop_btn.grid(row=1, column=3, sticky="ew", padx=(0, 5), pady=(2, 0))

        progress_frame = ttk.Frame(frame)
        progress_frame.pack(fill=tk.X)
        level1_frame = ttk.Frame(progress_frame)
        level1_frame.pack(fill=tk.X, pady=(2, 4))
        ttk.Label(level1_frame, text="一级:").pack(side=tk.LEFT, padx=(0, 5))
        level1_progress = ttk.Progressbar(level1_frame, orient="horizontal", mode="determinate")
        level1_progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        level1_label = ttk.Label(level1_frame, text="0/0", width=8)
        level1_label.pack(side=tk.LEFT)

        level2_frame = ttk.Frame(progress_frame)
        level2_frame.pack(fill=tk.X, pady=(2, 4))
        ttk.Label(level2_frame, text="二级:").pack(side=tk.LEFT, padx=(0, 5))
        level2_progress = ttk.Progressbar(level2_frame, orient="horizontal", mode="determinate")
        level2_progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        level2_label = ttk.Label(level2_frame, text="0/0", width=8)
        level2_label.pack(side=tk.LEFT)

        return {
            "frame": frame,
            "status_var": status_var,
            "pause_btn": pause_btn,
            "stop_btn": stop_btn,
            "entry": entry,
            "level1_progress": level1_progress,
            "level1_label": level1_label,
            "level2_progress": level2_progress,
            "level2_label": level2_label,
            "layout_mode": layout_mode,
        }


    def update_progress(self, level, current, total):
        """更新当前会话的进度条"""
        state = self._get_active_state()
        session_id = state.session_id if state else None
        # 使用after确保UI更新在主线程执行
        self.after(0, self._do_update_progress, session_id, level, current, total)

    def start_distribute_files(self):
        """在新线程中启动文件分配功能，避免UI阻塞"""
        self.log_ui("开始执行文件分配...")
        btn = getattr(self, "distribute_btn", None)
        if btn:
            btn.config(state=tk.DISABLED)
        threading.Thread(target=self.copy_matched_excel_files, daemon=True).start()

    def log_ui(self, message):
        # 使用lambda捕获message，通过after在主线程执行
        self.after(0, lambda msg=message: self._thread_safe_log(msg))

    def _thread_safe_log(self, message):
        """实际执行日志缓冲的方法（仅在主线程调用）"""
        if not hasattr(self, "output_text") or getattr(self, "output_text", None) is None:
            self.log_console(message)
            return
        self.log_buffer.append(f"[{time.strftime('%H:%M:%S')}] {message}\n")
        if len(self.log_buffer) >= self.log_batch_size and not self.log_updating:
            self.log_updating = True
            self.after(50, self._batch_update_log)

    def log_console(self, message):
        """在控制台显示详细调试信息"""
        print(f"[{time.strftime('%H:%M:%S')}] {message}")

    def _batch_update_log(self):
        """批量更新日志到UI（精确判断是否需要自动滚动）"""
        if not hasattr(self, "output_text") or getattr(self, "output_text", None) is None:
            self.log_buffer.clear()
            self.log_updating = False
            return
        if not self.log_buffer:
            self.log_updating = False
            return

        # 1. 批量处理缓冲区日志
        batch_log = ''.join(self.log_buffer)
        self.log_buffer.clear()

        # 2. 优化UI操作：减少状态切换次数
        self.output_text.config(state=tk.NORMAL)

        # 3. 记录插入前的滚动位置（用于判断是否需要自动滚动）
        # yview()返回一个元组：(top, bottom)，表示当前可见区域的上下边界（0.0-1.0）
        before_scroll = self.output_text.yview()
        is_scrolled_to_end = before_scroll[1] >= 0.99  # 允许微小误差（接近底部即视为需要自动滚动）

        # 4. 插入新日志
        self.output_text.insert(tk.END, batch_log)

        # 5. 限制最大行数，避免文本过多导致卡顿
        current_lines = int(self.output_text.index('end-1c').split('.')[0])
        if current_lines > self.log_max_lines:
            delete_lines = current_lines - self.log_max_lines
            self.output_text.delete(1.0, f"{delete_lines + 1}.0")

        # 6. 智能滚动：仅当用户未手动滚动（原位置在底部附近）时，才自动滚到底部
        if is_scrolled_to_end:
            self.output_text.see(tk.END)

        # 7. 恢复状态
        self.output_text.config(state=tk.DISABLED)
        self.log_updating = False

    def get_yesterday_date(self):
        """获取昨天的日期，格式为YYYY-MM-DD"""
        yesterday = datetime.now() - timedelta(days=1)
        return yesterday.strftime("%Y-%m-%d")

    def open_browser(self, session_id):
        """为指定会话打开浏览器"""
        state = self._ensure_session_state(session_id)
        if state.browser:
            self.log_ui(f"窗口{session_id + 1} 浏览器已打开")
            return

        self._set_active_session(session_id)
        self.set_session_status(session_id, "初始化中")
        self._run_in_session_thread(state, self._start_browser_session)

    def get_chrome_version(self):
        """获取当前安装的 Chrome 浏览器版本"""
        try:
            # 方法1：从注册表读取（最可靠）
            import winreg
            try:
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Google\Chrome\BLBeacon")
                version, _ = winreg.QueryValueEx(key, "version")
                winreg.CloseKey(key)
                # 提取主版本号，例如 "143.0.7499.147" -> "143"
                match = re.search(r'(\d+)\.', version)
                if match:
                    return match.group(1)
            except:
                pass
            
            # 方法2：尝试从系统注册表读取
            try:
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall\Google Chrome")
                version, _ = winreg.QueryValueEx(key, "version")
                winreg.CloseKey(key)
                match = re.search(r'(\d+)\.', version)
                if match:
                    return match.group(1)
            except:
                pass
            
            # 方法3：执行 chrome.exe --version
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            ]
            
            for chrome_path in chrome_paths:
                if os.path.exists(chrome_path):
                    try:
                        result = subprocess.run([chrome_path, "--version"], 
                                              capture_output=True, text=True, timeout=5)
                        version_str = result.stdout.strip()
                        # 提取主版本号
                        match = re.search(r'(\d+)\.', version_str)
                        if match:
                            return match.group(1)
                    except:
                        continue
                        
        except Exception as e:
            self.log_console(f"获取Chrome版本失败: {e}")
        return None

    def get_or_update_chromedriver(self):
        """智能获取或更新 ChromeDriver（仅在版本不匹配时下载）"""
        # 使用锁保护，防止多线程并发访问时的竞态条件
        with self.lock:
            if self.cached_driver_path and os.path.exists(self.cached_driver_path):
                return self.cached_driver_path
            
            # 固定的 ChromeDriver 目录
            driver_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chromedriver-win64")
            driver_path = os.path.join(driver_dir, "chromedriver.exe")
            os.makedirs(driver_dir, exist_ok=True)
            
            chrome_version = self.get_chrome_version()
            need_download = True
            
            # 检查现有驱动版本
            if os.path.exists(driver_path):
                try:
                    result = subprocess.run([driver_path, "--version"], 
                                          capture_output=True, text=True, timeout=5)
                    driver_version_str = result.stdout.strip()
                    # 提取驱动主版本号
                    match = re.search(r'(\d+)\.', driver_version_str)
                    if match:
                        driver_version = match.group(1)
                        if chrome_version:
                            # Chrome 版本检测成功，比较版本
                            if driver_version == chrome_version:
                                self.log_ui(f"✅ ChromeDriver 版本 {driver_version} 匹配 Chrome {chrome_version}，无需下载")
                                need_download = False
                            else:
                                self.log_ui(f"⚠️ 版本不匹配：Chrome {chrome_version} vs Driver {driver_version}，需要更新")
                        else:
                            # Chrome 版本检测失败，如果本地有驱动就直接使用
                            self.log_ui(f"⚠️ 无法检测 Chrome 版本，将使用现有 ChromeDriver {driver_version}")
                            need_download = False
                except Exception as e:
                    self.log_console(f"检查驱动版本失败: {e}")
                    # 如果驱动文件存在但无法执行，可能需要重新下载
                    if chrome_version:
                        self.log_ui(f"本地驱动异常，将重新下载 Chrome {chrome_version} 对应的驱动")
                    else:
                        self.log_ui(f"⚠️ Chrome 版本检测失败且本地驱动异常，尝试下载最新驱动")
            else:
                if chrome_version:
                    self.log_ui(f"本地未找到 ChromeDriver，将下载 Chrome {chrome_version} 对应的驱动")
                else:
                    self.log_ui(f"⚠️ Chrome 版本检测失败且无本地驱动，尝试下载最新驱动")
            
            if need_download:
                try:
                    self.log_ui(f"⬇️ 正在下载 Chrome {chrome_version} 对应的 ChromeDriver...")
                    # 使用 webdriver_manager 下载
                    downloaded_path = ChromeDriverManager().install()
                    
                    # 复制到固定目录
                    if os.path.exists(downloaded_path):
                        shutil.copy2(downloaded_path, driver_path)
                        self.log_ui(f"✅ ChromeDriver 已更新到: {driver_dir}")
                    else:
                        self.log_ui(f"❌ 下载的驱动路径无效: {downloaded_path}")
                        return downloaded_path
                except Exception as e:
                    self.log_ui(f"❌ 下载 ChromeDriver 失败: {e}")
                    # 如果下载失败但本地有旧版本，尝试使用旧版本
                    if os.path.exists(driver_path):
                        self.log_ui("⚠️ 将使用现有版本的 ChromeDriver")
                    else:
                        raise
            
            self.cached_driver_path = driver_path
            return driver_path


    def _start_browser_session(self):
        """在当前线程绑定的会话中启动浏览器"""
        state = self._get_active_state()
        try:
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument(f"--remote-debugging-port={state.port}")
            chrome_options.add_argument(f"--user-data-dir={state.profile_path}")

            # 智能获取或更新 ChromeDriver（仅在版本不匹配时下载）
            driver_path = self.get_or_update_chromedriver()
            service = Service(driver_path)
            self.browser = webdriver.Chrome(service=service, options=chrome_options)
            self.browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                "source": """
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    })
                """
            })

            self.browser.get("https://sycm.taobao.com/")
            self.log_session(f"浏览器启动成功（端口 {state.port}）")
            self.log_session("请在新窗口中完成登录")
            self.after(0, lambda sid=state.session_id: self.set_session_status(sid, "浏览器已启动，等待登录"))

            # 将第一个就绪的会话设置为当前会话
            if self.active_session_id < 0:
                self.after(0, lambda sid=state.session_id: self._set_active_session(sid))

        except WebDriverException as e:
            error_msg = str(e)
            self.log_session(f"浏览器启动失败（端口 {state.port}）: {error_msg}")
            self.log_console(f"浏览器启动详细错误: {traceback.format_exc()}")

            lower_msg = error_msg.lower()
            if "address already in use" in lower_msg or "debugging port" in lower_msg or "bind()" in lower_msg:
                self.log_session(f"端口 {state.port} 可能被占用，请确认无其他实例占用后重试")
            elif "user data directory" in lower_msg or "devtoolsactiveport" in lower_msg:
                self.log_session(f"检测到用户数据目录被占用，请清理 {state.profile_path} 后重试")
            self.after(0, lambda sid=state.session_id: self.set_session_status(sid, "浏览器启动失败"))
        except Exception as exc:  # pylint: disable=broad-except
            self.log_ui(f"启动过程发生意外错误: {exc}")
            self.log_console(traceback.format_exc())
            self.after(0, lambda sid=state.session_id: self.set_session_status(sid, "启动失败"))

    def open_interface(self, session_id):
        """为指定窗口打开工作界面"""
        state = self.session_states.get(session_id)
        if not state or not state.browser:
            self.log_ui(f"窗口{session_id + 1} 请先打开浏览器并登录")
            return

        self._set_active_session(session_id)
        self.set_session_status(session_id, "初始化中")
        self._run_in_session_thread(state, self._prepare_work_interface)

    def get_monday_date(self):
        """获取最近的星期一日期（若当天是星期一则返回当天，否则返回上一个星期一）"""
        today = datetime.now()
        # weekday()返回值：0=星期一，1=星期二，...，6=星期日
        weekday = today.weekday()

        # 计算与最近星期一的差值：若当天是星期一（0），差值为0；否则差值为“当前星期几的数值”
        days_to_monday = weekday  # 例如：星期二（1）→ 差1天，星期日（6）→ 差6天
        monday = today - timedelta(days=days_to_monday)

        return monday.strftime("%Y-%m-%d")

    def _prepare_work_interface(self):
        state = self._get_active_state()
        try:
            # 获取最近的星期一日期
            monday_date = self.get_monday_date()
            self.log_session(f"使用日期: {monday_date}（最近的星期一）")

            # 构建带昨天日期的URL
            yesterday = self.get_yesterday_date()
            target_url = f"https://sycm.taobao.com/mc/free/search_rank?dateRange={yesterday}%7C{yesterday}&dateType=day&cateId=11&cateFlag=1&parentCateId=50007216"
            self.browser.get(target_url)
            if not self.wait_for_work_page_ready():
                self.after(0, lambda sid=state.session_id: self.set_session_status(sid, "初始化失败"))
                return

            # 新增：界面打开成功后，设置状态变量为True
            self.interface_opened = True

            # 自动选择50条/页的选项
            self.select_50_via_working_method()

            # 打开类目选择器并获取一级类目
            if self.open_category_picker():
                self.get_categories_by_level(1)
                self.log_session("工作界面准备完成，请输入类目操作指令")
                self.after(0, lambda sid=state.session_id: self.set_session_status(sid, "初始化完成"))
            else:
                self.log_session("打开类目选择器失败，请手动操作或重试")
                self.after(0, lambda sid=state.session_id: self.set_session_status(sid, "初始化失败"))

        except Exception as e:
            self.log_session(f"准备工作界面出错: {str(e)}")
            self.log_console(f"准备工作错误详情: {traceback.format_exc()}")
            self.after(0, lambda sid=state.session_id: self.set_session_status(sid, "初始化失败"))

    def wait_for_work_page_ready(self, timeout=25):
        """等待工作页面加载完成"""
        if not self.browser:
            self.log_ui("浏览器未初始化，无法检测页面状态")
            return False
        try:
            WebDriverWait(self.browser, timeout).until(
                lambda driver: driver.execute_script("return document.readyState") == "complete"
            )
            WebDriverWait(self.browser, timeout).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "div.oui-pro-common-picker.isready, div.category-picker"))
            )
            self.log_session("页面加载完成")
            return True
        except TimeoutException:
            self.log_session("等待页面加载超时，请检查网络或重新尝试")
            return False
        except Exception as exc:  # pylint: disable=broad-except
            self.log_session(f"等待页面加载时出错: {exc}")
            self.log_console(traceback.format_exc())
            return False

    def select_50_via_working_method(self):
        """使用已验证成功的方法选择50条/页"""
        try:
            self.log_ui("开始设置显示数量为50...")

            # 使用已验证成功的选择器
            selectors = [
                "ant-select-sm.oui-select.oui-page-size-select",
                "ant-select-sm.oui-select.oui-page-size-select.ant-select.ant-select-enabled"
            ]

            success = False
            for selector in selectors:
                try:
                    # 构建CSS选择器
                    css_selector = "." + selector.replace(".", ".")
                    self.log_console(f"使用选择器: {css_selector}")

                    # 等待元素出现并点击
                    container = WebDriverWait(self.browser, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, css_selector))
                    )
                    container.click()
                    time.sleep(1)  # 等待下拉框展开

                    # 查找并点击50选项
                    option = WebDriverWait(self.browser, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//li[text()='50']"))
                    )
                    option.click()
                    time.sleep(1)

                    self.log_ui("成功将显示数量设置为50")
                    success = True
                    break

                except NoSuchElementException:
                    self.log_console(f"选择器 {selector} 未找到元素")
                except Exception as e:
                    self.log_console(f"选择器 {selector} 操作失败: {str(e)}")

            if not success:
                self.log_ui("选择50条/页失败，请手动设置")

        except Exception as e:
            self.log_ui(f"设置显示数量时出错: {str(e)}")
            self.log_console(f"执行选择时发生错误: {str(e)}")

    def open_category_picker(self):
        """打开类目选择器"""
        if not self.browser:
            return False

        try:
            self.log_ui("尝试打开类目选择器...")

            # 等待选择器容器出现
            WebDriverWait(self.browser, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.oui-pro-common-picker, div.category-picker"))
            )

            # 使用JavaScript打开选择器
            result = self.browser.execute_script("""
                var picker = document.querySelector('div.oui-pro-common-picker.isready');
                if (!picker) {
                    picker = document.querySelector('div.category-picker');
                }

                if (!picker) {
                    return '未找到类目选择器容器';
                }

                picker.classList.add('open');
                var event = new Event('change', {bubbles: true});
                picker.dispatchEvent(event);
                return 'success';
            """)

            if result == 'success':
                time.sleep(2)
                self.log_ui("类目选择器已打开")
                return True
            else:
                self.log_ui(f"打开类目选择器失败: {result}")
                return False

        except Exception as e:
            self.log_ui(f"打开类目选择器时出错: {str(e)}")
            return False

    def _do_update_progress(self, session_id, level, current, total):
        """实际更新进度条的方法"""
        panel = getattr(self, "session_panels", {}).get(session_id)
        if not panel:
            return

        if total == 0:
            percentage = 0
        else:
            percentage = (current / total) * 100

        if level == 1:
            panel["level1_progress"]["value"] = percentage
            panel["level1_label"]["text"] = f"{current}/{total}"
            state = self.session_states.get(session_id)
            if state:
                state.level1_current = current
                state.level1_total = total
        elif level == 2:
            panel["level2_progress"]["value"] = percentage
            panel["level2_label"]["text"] = f"{current}/{total}"
            state = self.session_states.get(session_id)
            if state:
                state.level2_current = current
                state.level2_total = total

    def reset_progress(self, level=None):
        """重置进度条（level为None时重置所有）"""
        state = self._get_active_state()
        session_id = state.session_id if state else None
        if level is None or level == 1:
            self.after(0, self._do_update_progress, session_id, 1, 0, 0)
        if level is None or level == 2:
            self.after(0, self._do_update_progress, session_id, 2, 0, 0)

    def detect_target_iframe(self, scene="数据提取前"):
        """检测目标iframe元素，新增自动暂停逻辑"""
        if not self.browser:
            self.log_ui(f"[{scene}] 浏览器未初始化，无法检测iframe")
            return False

        try:
            self.log_ui(f"[{scene}] 开始检测目标iframe元素...")
            iframe_element = WebDriverWait(self.browser, 1).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    '//iframe[@id="baxia-dialog-content" '
                    'and @frameborder="none" '
                    'and contains(@src, "https://sycm.taobao.com:443//mc/mq/mkt/keyword/rank.json") '
                    'and contains(@src, "action=captcha") '
                    'and contains(@src, "pureCaptcha=")]'
                ))
            )

            src = iframe_element.get_attribute("src")
            self.log_ui(f"[{scene}] ✅ 检测到目标iframe元素!")
            self.log_ui(f"[{scene}]   id: {iframe_element.get_attribute('id')}")
            self.log_ui(f"[{scene}]   src(部分): {src[:100]}...")

            # 【新增】检测到iframe后自动暂停提取流程
            self.log_ui(f"[{scene}] ⏸️ 检测到目标iframe，自动暂停数据提取")
            self.toggle_pause()  # 调用暂停方法

            return True

        except TimeoutException:
            self.log_ui(f"[{scene}] ❌ 未找到目标iframe元素（超时）")
            return False
        except NoSuchElementException:
            self.log_ui(f"[{scene}] ❌ 未找到目标iframe元素（元素不存在）")
            return False
        except Exception as e:
            self.log_ui(f"[{scene}] 检测iframe时发生错误: {str(e)}")
            return False

    def get_categories_by_level(self, level):
        """获取指定层级的类目（优化三级类目加载逻辑）"""
        if not self.browser:
            return []

        try:
            self.log_ui(f"正在获取第{level}级类目...")

            # 等待类目容器加载完成（最长等待15秒）
            WebDriverWait(self.browser, 15).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, f"ul.tree-menu.common-menu.tree-scroll-menu-level-{level}"))
            )

            # 针对三级类目增加滚动加载逻辑（解决数量过多时元素未渲染问题）
            if level == 3:
                self.log_ui("检测到三级类目，执行滚动加载确保所有元素可见...")
                # 1. 先滚动到容器底部，触发懒加载
                self.browser.execute_script("""
                    var level = arguments[0];
                    var levelClass = 'tree-scroll-menu-level-' + level;
                    var containerSelector = 'ul.tree-menu.common-menu.' + levelClass;
                    var container = document.querySelector(containerSelector);

                    if (container) {
                        // 记录原始滚动位置
                        var originalScrollTop = container.scrollTop;
                        // 滚动到底部
                        container.scrollTop = container.scrollHeight;
                        // 等待2秒确保元素加载
                        setTimeout(() => {}, 2000);
                        // 滚回原始位置
                        container.scrollTop = originalScrollTop;
                    }
                """, level)
                # 等待滚动加载完成
                time.sleep(2)

            # 使用JavaScript获取类目数据
            categories = self.browser.execute_script("""
                var level = arguments[0];
                var levelClass = 'tree-scroll-menu-level-' + level;
                var containerSelector = 'ul.tree-menu.common-menu.' + levelClass;

                var container = document.querySelector(containerSelector);
                if (!container) {
                    return {status: 'error', message: '未找到第' + level + '级类目容器'};
                }

                // 再次滚动确保所有元素已加载（针对三级类目）
                if (level === 3) {
                    container.scrollTop = 0;
                    setTimeout(() => {}, 500);
                }

                var items = container.querySelectorAll('li.tree-item.common-item');
                var results = [];

                for (var i = 0; i < items.length; i++) {
                    // 获取类目名称（优先使用title属性，无则用文本内容）
                    var name = items[i].getAttribute('title') || items[i].textContent.trim();
                    if (name) {
                        var hasChildren = false;

                        // 针对二级类目，检查是否有三级类目图标
                        if (level === 2) {
                            var icon = items[i].querySelector(
                                'i.anticon.anticon-angle-right.oui-canary-icon.oui-canary-icon-angle-right.sub-tree-icon.sub-common-icon'
                            );
                            hasChildren = !!icon; // 存在图标则表示有子类目
                        }

                        results.push({
                            name: name,
                            index: i + 1,  // 索引从1开始
                            has_children: hasChildren
                        });
                    }
                }

                return {
                    status: 'success',
                    data: results
                };
            """, level)

            # 处理返回结果
            if categories['status'] != 'success':
                self.log_ui(f"获取类目失败: {categories['message']}")
                return []

            # 存储并显示类目数据
            self.categories[level] = categories['data']
            self.output_categories(level)  # 打印类目列表到日志

            return categories['data']

        except TimeoutException:
            self.log_ui(f"获取第{level}级类目超时（容器未加载）")
            return []
        except Exception as e:
            self.log_ui(f"获取第{level}级类目时出错: {str(e)}")
            self.log_console(f"获取类目详细错误: {traceback.format_exc()}")  # 控制台打印详细堆栈
            return []

    def output_categories(self, level):
        """在UI中显示类目列表"""
        self.log_ui(f"\n第{level}级类目列表:")
        for item in self.categories[level]:
            if level == 2 and item['has_children']:
                self.log_ui(f"{item['index']}. {item['name']} (有三级类目)")
            else:
                self.log_ui(f"{item['index']}. {item['name']}")
        self.log_ui("")  # 空行分隔

    def process_input(self, session_id, event=None):
        """处理指定会话的输入指令"""
        panel = getattr(self, "session_panels", {}).get(session_id)
        if not panel:
            return

        entry_widget = panel["entry"]
        input_text = entry_widget.get().strip()
        entry_widget.delete(0, tk.END)

        if not input_text:
            self.log_ui("请输入类目操作指令")
            return

        state = self.session_states.get(session_id)
        if not state or not state.browser:
            self.log_ui(f"窗口{session_id + 1} 请先打开浏览器并完成登录")
            return
        if not state.interface_opened:
            self.log_ui(f"窗口{session_id + 1} 请先打开工作界面")
            return
        if state.processing:
            self.log_ui(f"窗口{session_id + 1} 正在执行任务，请稍候")
            return

        self._set_active_session(session_id)

        if "-" in input_text:
            try:
                start_idx, end_idx = map(int, input_text.split("-"))
            except ValueError:
                self.log_ui("范围格式错误，请使用 5-8 形式")
                return

            if start_idx < 1 or end_idx < start_idx or end_idx > len(self.categories[1]):
                self.log_ui(f"无效范围，请输入 1-{len(self.categories[1])} 之间的数字")
                return

            if self.current_level != 1:
                self.log_ui("范围提取仅支持一级类目")
                return

            self.log_session(f"开始处理范围 {start_idx}-{end_idx}")
            self._start_processing_task(state, self.process_level1_range, start_idx, end_idx)
            return

        if input_text == "0":
            if self.current_level != 1:
                self.log_ui("仅在一级类目下可使用提取全部")
                return

            total = len(self.categories[1])
            self.log_session(f"开始提取全部一级类目，共 {total} 个")
            self._start_processing_task(state, self.process_level1_range, 1, total)
            return

        try:
            if input_text.endswith('.'):
                level_index = int(input_text[:-1])
                action = "select"
            else:
                level_index = int(input_text)
                action = "extract"
        except ValueError:
            self.log_ui("指令格式错误，可使用 0、数字 或 范围 5-8")
            return

        if level_index < 1 or level_index > len(self.categories[self.current_level]):
            self.log_ui(f"无效编号，请输入 1-{len(self.categories[self.current_level])} 之间的数字")
            return

        selected = next((item for item in self.categories[self.current_level] if item['index'] == level_index), None)
        if not selected:
            self.log_ui(f"未找到编号为 {level_index} 的类目")
            return

        self.selected_categories[self.current_level] = selected
        self.log_session(f"选中第{self.current_level}级类目：{selected['name']}")

        if not self.click_category(self.current_level, level_index, selected['name']):
            return

        if action == "select":
            self.log_session(f"已切换到第{self.current_level + 1}级类目")
            self.current_level += 1
            self.get_categories_by_level(self.current_level)
        else:
            self.log_session(f"开始提取第{self.current_level}级类目 {selected['name']} 的数据")
            self.current_excel_root = self.create_root_excel_file(level_index, selected['name'])
            self.open_excel_file()
            self._start_processing_task(state, self.process_category_data)

    def click_category(self, level, index, name):
        """在类目树中点击指定的类目"""
        if not self.browser:
            self.log_session("浏览器尚未就绪，无法点击类目")
            return False
        try:
            selector = f"ul.tree-scroll-menu-level-{level} li.tree-item.common-item"
            result = self.browser.execute_script("""
                var sel = arguments[0];
                var targetIndex = arguments[1];
                var targetName = arguments[2];

                var items = document.querySelectorAll(sel);
                if (!items || items.length === 0) {
                    return '未找到类目列表元素';
                }
                if (targetIndex < 1 || targetIndex > items.length) {
                    return `索引超出范围(共${items.length}项)`;
                }

                var target = items[targetIndex - 1];
                if (!target) {
                    return '无法定位到目标元素';
                }
                var title = target.getAttribute('title') || '';
                var text = target.textContent ? target.textContent.trim() : '';
                var actual = title || text;
                if (actual && targetName && actual.trim() !== targetName.trim()) {
                    return `名称不匹配: ${actual}`;
                }
                try {
                    target.click();
                } catch (e) {
                    var evt = new MouseEvent('click', {bubbles: true, cancelable: true, view: window});
                    target.dispatchEvent(evt);
                }
                return 'success';
            """, selector, index, name)

            if result == 'success':
                self.log_session(f"成功点击第{level}级类目（序号{index}）: {name}")
                time.sleep(1)
                return True

            self.log_session(f"点击类目失败: {result}")
            self.get_categories_by_level(level)
            return False

        except Exception as exc:  # pylint: disable=broad-except
            self.log_session(f"点击类目出错: {exc}")
            self.log_console(traceback.format_exc())
            return False
    def process_level1_range(self, start, end):
        """处理一级类目的范围提取（从start到end）"""
        try:
            # 获取所有一级类目
            level1_cats = self.categories[1].copy()
            range_cats = [cat for cat in level1_cats if start <= cat['index'] <= end]
            total_range = len(range_cats)  # 范围内总类目数（含排除项）

            # 筛选范围内的类目
            total_process = sum(1 for cat in range_cats if cat['index'] not in self.exclude_level1_serials)
            self.log_ui(
                f"范围[{start}-{end}]内共{total_range}个一级类目，其中{total_range - total_process}个需排除，实际处理{total_process}个")
            self.update_progress(1, 0, total_process)  # 进度条最大值设为“实际处理数”

            # 点击第一个类目初始化
            first_valid_cat = None
            for cat in range_cats:
                if cat['index'] not in self.exclude_level1_serials:
                    first_valid_cat = cat
                    break
            if first_valid_cat and self.browser:
                if not self.open_category_picker():
                    self.log_ui("无法打开类目选择器，程序退出")
                    return
                self.log_ui(
                    f"首次点击范围类目中的有效项（原始序号{first_valid_cat['index']}）: {first_valid_cat['name']}")
                if not self.click_category(1, first_valid_cat['index'], first_valid_cat['name']):
                    self.log_ui("首次点击失败，无法继续")
                    return
                self.current_level1_index = first_valid_cat['index']
                time.sleep(2)
                self.current_level = 2
                self.get_categories_by_level(2)

            # 遍历范围内的一级类目
            processed_count = 0  # 实际处理数（用于进度条）
            for level1_cat in range_cats:
                level1_idx = level1_cat['index']  # 原始序号（不修改）
                level1_name = level1_cat['name']

                # 1. 检查是否为排除项：是则跳过
                if level1_idx in self.exclude_level1_serials:
                    self.log_ui(f"\n===== 跳过范围中排除的类目（原始序号{level1_idx}）: {level1_name} =====")
                    continue

                # 2. 检查是否需要停止
                if self.stop_event.is_set():
                    self.log_ui("已结束范围一级类目处理")
                    break

                # 3. 更新进度（实际处理数+1）
                processed_count += 1
                self.update_progress(1, processed_count, total_process)
                self.log_ui(f"\n===== 开始处理范围类目中的有效项（原始序号{level1_idx}）: {level1_name} =====")

                # 4. 后续原有处理逻辑不变
                if not self.open_category_picker():
                    self.log_ui("尝试重新打开类目选择器")
                    if not self.open_category_picker():
                        self.log_ui(f"跳过类目（原始序号{level1_idx}）: 无法打开选择器")
                        continue

                # 切换到当前一级类目
                if level1_idx != self.current_level1_index:
                    self.log_ui(f"切换到一级类目（原始序号{level1_idx}）: {level1_name}")
                    if not self.click_category(1, level1_idx, level1_name):
                        self.log_ui(f"跳过类目（原始序号{level1_idx}）: 点击失败")
                        continue
                    self.current_level1_index = level1_idx
                    time.sleep(2)
                    self.current_level = 2
                    self.get_categories_by_level(2)

                # 创建Excel文件（用原始序号命名）
                self.selected_categories[1] = level1_cat
                self.current_excel_root = self.create_root_excel_file(level1_idx, level1_name)
                self.open_excel_file()
                self._reset_collected_data()

                # 处理当前一级类目下的二级类目
                self.process_level1_category(level1_idx, total_level1=total_process)

                if self.stop_event.is_set():
                    self._reset_collected_data()
                else:
                    self._flush_collected_data()

                # 检查停止状态
                if self.stop_event.is_set():
                    self.log_ui("已结束范围一级类目处理")
                    break

        except Exception as e:
            self.log_ui(f"处理范围一级类目时出错: {str(e)}")
            self.log_console(f"范围处理错误详情: {traceback.format_exc()}")
        finally:
            # 恢复状态并重置到一级类目
            with self.lock:
                self.processing = False
                self.paused = False
                self.stop_event.clear()
                self.selected_categories = {1: None, 2: None, 3: None}
                self.current_level1_index = None

            self.log_ui(f"第{start}到第{end}个一级类目数据提取完成")

            # 重置到一级类目
            if self.browser and not self.stop_event.is_set():
                try:
                    if self.open_category_picker():
                        self.current_level = 1
                        self.get_categories_by_level(1)
                        self.log_ui("已重置到一级类目，请输入新的操作指令")
                    else:
                        self.log_ui("重置失败，请点击'打开界面'重新加载一级类目")
                except Exception as e:
                    self.log_ui(f"重置到一级类目时出错: {str(e)}")

    def process_all_level1_categories(self):
        """处理所有一级类目"""
        try:
            # 获取所有一级类目
            level1_cats = self.categories[1].copy()
            total_all = len(level1_cats)
            total_process = sum(1 for cat in level1_cats if cat['index'] not in self.exclude_level1_serials)
            self.log_ui(
                f"所有一级类目共{total_all}个，其中{len(self.exclude_level1_serials)}个需排除，实际处理{total_process}个")
            self.update_progress(1, 0, total_process)  # 进度条最大值设为“实际处理数”

            # 点击第一个一级类目
            first_valid_cat = None
            for cat in level1_cats:
                if cat['index'] not in self.exclude_level1_serials:
                    first_valid_cat = cat
                    break
            if first_valid_cat and self.browser:
                if not self.open_category_picker():
                    self.log_ui("无法打开类目选择器，程序退出")
                    return
                # 点击第一个有效类目初始化
                self.log_ui(f"首次点击一级类目（原始序号{first_valid_cat['index']}）: {first_valid_cat['name']}")
                if not self.click_category(1, first_valid_cat['index'], first_valid_cat['name']):
                    self.log_ui("首次点击失败，无法继续")
                    return
                self.current_level1_index = first_valid_cat['index']
                time.sleep(2)
                self.current_level = 2
                self.get_categories_by_level(2)

            # 遍历所有一级类目
            processed_count = 0  # 记录实际处理的数量（用于进度条）
            for level1_cat in level1_cats:
                level1_idx = level1_cat['index']  # 原始序号（不修改）
                level1_name = level1_cat['name']

                # 1. 检查是否为排除项：是则跳过
                if level1_idx in self.exclude_level1_serials:
                    self.log_ui(f"\n===== 跳过排除的一级类目（原始序号{level1_idx}）: {level1_name} =====")
                    continue

                # 2. 检查是否需要停止
                if self.stop_event.is_set():
                    self.log_ui("已结束全量一级类目处理")
                    break

                # 3. 更新进度（实际处理数+1）
                processed_count += 1
                self.update_progress(1, processed_count, total_process)
                self.log_ui(f"\n===== 开始处理一级类目（原始序号{level1_idx}）: {level1_name} =====")

                # 4. 后续原有处理逻辑（点击类目、处理二级类目等）不变
                if not self.open_category_picker():
                    self.log_ui("尝试重新打开类目选择器")
                    if not self.open_category_picker():
                        self.log_ui(f"跳过类目（原始序号{level1_idx}）: 无法打开选择器")
                        continue

                # 切换到当前一级类目（如果不是当前选中的）
                if level1_idx != self.current_level1_index:
                    self.log_ui(f"切换到一级类目（原始序号{level1_idx}）: {level1_name}")
                    if not self.click_category(1, level1_idx, level1_name):
                        self.log_ui(f"跳过类目（原始序号{level1_idx}）: 点击失败")
                        continue
                    self.current_level1_index = level1_idx
                    time.sleep(2)
                    self.current_level = 2
                    self.get_categories_by_level(2)

                # 创建Excel文件（用原始序号命名）
                self.selected_categories[1] = level1_cat
                self.current_excel_root = self.create_root_excel_file(level1_idx, level1_name)  # 文件名用原始序号
                self.open_excel_file()
                self._reset_collected_data()

                # 处理当前一级类目下的二级类目
                self.process_level1_category(level1_idx, total_level1=total_process)

                if self.stop_event.is_set():
                    self._reset_collected_data()
                else:
                    self._flush_collected_data()

                # 再次检查停止状态
                if self.stop_event.is_set():
                    self.log_ui("已结束全量一级类目处理")
                    break

        except Exception as e:
            self.log_ui(f"处理全量一级类目时出错: {str(e)}")
        finally:
            # 重置进度条
            self.reset_progress()
            # 恢复状态并重置到一级类目
            with self.lock:
                self.processing = False
                self.paused = False
                self.stop_event.clear()
                self.selected_categories = {1: None, 2: None, 3: None}
                self.current_level1_index = None

            self.log_ui("全量一级类目数据提取完成，可以重新选择类目开始")

            # 只在未被手动停止的情况下执行一次重置
            if self.browser and not self.stop_event.is_set():
                try:
                    if self.open_category_picker():
                        self.current_level = 1
                        self.get_categories_by_level(1)
                        self.log_ui("已重置到一级类目，请输入新的操作指令")
                    else:
                        self.log_ui("重置失败，请点击'打开界面'重新加载一级类目")
                except Exception as e:
                    self.log_ui(f"重置到一级类目时出错: {str(e)}")

    def process_assigned_categories(self, assigned_indices):
        """处理指定编号的一级类目集合"""
        try:
            if not assigned_indices:
                self.log_ui("未提供需要处理的类目编号")
                return

            order_map = {idx: pos for pos, idx in enumerate(assigned_indices)}
            level1_cats = [
                cat for cat in self.categories[1]
                if cat['index'] in order_map and cat['index'] not in self.exclude_level1_serials
            ]
            if not level1_cats:
                self.log_ui("当前列表中未找到指定的一级类目")
                return

            level1_cats.sort(key=lambda cat: order_map[cat['index']])
            total_process = len(level1_cats)
            self.log_ui(f"开始处理指定的{total_process}个一级类目: {assigned_indices}")
            self.update_progress(1, 0, total_process)

            first_valid_cat = level1_cats[0]
            if self.browser:
                if not self.open_category_picker():
                    self.log_ui("无法打开类目选择器，终止处理")
                    return
                if not self.click_category(1, first_valid_cat['index'], first_valid_cat['name']):
                    self.log_ui("首次点击类目失败，终止处理")
                    return
                self.current_level1_index = first_valid_cat['index']
                time.sleep(2)
                self.current_level = 2
                self.get_categories_by_level(2)

            processed_count = 0
            for level1_cat in level1_cats:
                if self.stop_event.is_set():
                    self.log_ui("检测到停止指令，结束当前窗口任务")
                    break

                level1_idx = level1_cat['index']
                level1_name = level1_cat['name']

                processed_count += 1
                self.update_progress(1, processed_count, total_process)
                self.log_ui(f"\n===== 开始处理指定类目（原始序号{level1_idx}）: {level1_name} =====")

                if not self.open_category_picker():
                    self.log_ui("类目选择器加载失败，尝试重新打开")
                    if not self.open_category_picker():
                        self.log_ui("仍无法打开类目选择器，跳过该类目")
                        continue

                if level1_idx != self.current_level1_index:
                    self.log_ui(f"切换到一级类目（原始序号{level1_idx}）: {level1_name}")
                    if not self.click_category(1, level1_idx, level1_name):
                        self.log_ui("切换类目失败，跳过该类目")
                        continue
                    self.current_level1_index = level1_idx
                    time.sleep(2)
                    self.current_level = 2
                    self.get_categories_by_level(2)

                self.selected_categories[1] = level1_cat
                self.current_excel_root = self.create_root_excel_file(level1_idx, level1_name)
                self.open_excel_file()
                self._reset_collected_data()

                self.process_level1_category(level1_idx, total_level1=total_process)

                if self.stop_event.is_set():
                    self._reset_collected_data()
                else:
                    self._flush_collected_data()

        except Exception as e:
            self.log_ui(f"处理指定一级类目时发生错误: {str(e)}")
            self.log_console(traceback.format_exc())
        finally:
            with self.lock:
                self.processing = False
                self.paused = False
                self.stop_event.clear()
                self.selected_categories = {1: None, 2: None, 3: None}
                self.current_level1_index = None

            if self.browser and not self.stop_event.is_set():
                try:
                    if self.open_category_picker():
                        self.current_level = 1
                        self.get_categories_by_level(1)
                        self.log_ui("指定类目处理完成，已回到一级类目列表")
                except Exception as e:
                    self.log_ui(f"重置一级类目时发生错误: {str(e)}")

    def process_category_data(self):
        """处理类目数据提取"""
        try:
            current_level = self.current_level

            if current_level == 1:
                # 获取当前选中的一级类目
                level1_cat = self.selected_categories.get(1)
                if not level1_cat:
                    self.log_ui("未选中有效的一级类目，无法处理")
                    return
                # 补充参数：当前一级类目索引、总一级类目数（此处按单个处理，总数量设为1）
                self._reset_collected_data()
                self.process_level1_category(
                    level1_idx=level1_cat['index'],  # 从选中的类目里取index
                    total_level1=1  # 单个类目处理时，总数量为1
                )
                if self.stop_event.is_set():
                    self._reset_collected_data()
                else:
                    self._flush_collected_data()
            elif current_level == 2:
                self._reset_collected_data()
                self.process_level2_category()
                if self.stop_event.is_set():
                    self._reset_collected_data()
                else:
                    self._flush_collected_data()

        except Exception as e:
            self.log_ui(f"处理类目数据时出错: {str(e)}")
        finally:
            # 恢复状态并重置到一级类目
            with self.lock:
                self.processing = False
                self.paused = False
                self.stop_event.clear()
                self.selected_categories = {1: None, 2: None, 3: None}

            self.log_ui("类目数据提取完成，可以重新选择类目开始")

    def check_first_page_for_low_value(self):
        """检查第一页是否有小于150的数据"""
        self.log_ui("检查第一页是否有小于150的搜索人气...")

        try:
            # 记录当前页码
            current_page = self.get_current_page_number()

            # 如果不在第一页，跳转到第一页
            if current_page != 1:
                self.log_ui("跳转到第一页")
                if not self.click_page_number(1):
                    self.log_ui("跳转至第一页失败，使用当前页进行检查")
                    return False

                time.sleep(1)

            # 提取第一页数据
            page_data, found_low_value = self.extract_data_from_page()

            return found_low_value

        except Exception as e:
            self.log_ui(f"检查第一页时出错: {str(e)}")
            return False

    def check_last_page_for_low_value(self):
        """检查最后一页是否有小于150的数据"""
        self.log_ui("检查最后一页是否有小于150的搜索人气...")

        try:
            # 记录当前页码
            current_page = self.get_current_page_number()

            # 获取最大页数
            max_page = self.get_max_page_number()

            if max_page and max_page > current_page:
                self.log_ui(f"检测到最大页数为: {max_page}，跳转到最后一页")

                # 跳转到最大页码
                if not self.click_page_number(max_page):
                    self.log_ui("跳转至最大页码失败，使用备选方案")
                    return False

                time.sleep(1)

                # 提取最后一页数据
                page_data, found_low_value = self.extract_data_from_page()

                # 回到第一页
                self.click_page_number(1)
                time.sleep(1)

                return found_low_value
            else:
                self.log_ui("当前已是最后一页或无法获取最大页数，直接检查当前页")
                # 直接检查当前页
                _, found_low_value = self.extract_data_from_page()
                return found_low_value

        except Exception as e:
            self.log_ui(f"检查最后一页时出错: {str(e)}")
            return False

    def process_level1_category(self, level1_idx, total_level1):
        """处理一级类目下的二级类目（强化停止检查）"""
        try:
            level1_cat = self.selected_categories[1]

            self.current_level = 2
            level2_cats = self.get_categories_by_level(2)
            total_level2 = len(level2_cats)

            # 初始化二级进度条
            self.update_progress(2, 0, total_level2)

            if not level2_cats:
                self.log_ui("该一级类目下没有二级类目，直接提取数据")
                if not self.collect_data_across_pages():
                    self.log_ui("当前类目数据为空")
                return

            for level2_idx, level2_cat in enumerate(level2_cats, 1):
                # 每次循环最开始检查停止指令（优先响应）
                if self.stop_event.is_set():
                    self.log_ui("检测到结束指令，立即停止处理后续二级类目")
                    return  # 直接退出循环，不处理下一个

                # 更新二级类目进度
                self.update_progress(2, level2_idx, total_level2)
                self.log_ui(f"\n===== 开始处理第{level2_idx}个二级类目: {level2_cat['name']} =====")

                if not self.click_category(2, level2_idx, level2_cat['name']):
                    continue

                # 处理二级类目之前再次检查
                if self.stop_event.is_set():
                    self.log_ui("检测到结束指令，不处理当前二级类目")
                    return

                if level2_cat['has_children']:
                    self.process_secondary_with_tertiary(level2_cat, level2_idx)
                else:
                    self.process_secondary_without_tertiary()


        finally:
            # 只有正常结束（非强制停止）才执行重置
            if not self.stop_event.is_set():
                self.current_level = 1
                if self.browser and self.open_category_picker():
                    self.get_categories_by_level(1)

            # 处理完一个一级类目后重置二级进度条
            if not self.stop_event.is_set():
                self.update_progress(2, 0, 0)

    def process_secondary_with_tertiary(self, level2_cat, level2_idx):
        """处理有三级类目的二级类目（增加密集检查点）"""
        # 检查1：开始处理前
        if not self.check_pause_state():
            return

        self.log_ui("检查当前二级类目是否有数据...")
        page_data, _ = self.extract_data_from_page()
        if not page_data:  # 如果数据为空
            self.log_ui("当前二级类目数据为空，跳过其下三级类目，直接处理下一个二级类目")
            return  # 直接返回，不处理三级类目

        self.log_ui("三级类目数据提取前，检测目标iframe...")
        self.detect_target_iframe()

        # 先检查第一页是否有<150的数据
        has_low_value_first = self.check_first_page_for_low_value()

        if has_low_value_first:
            # 检查2：提取数据前
            if not self.check_pause_state():
                return

            # 第一页有低价值数据：只提取当前页
            self.log_ui("第一页发现小于150的搜索人气，仅提取当前页数据")
            page_data, _ = self.extract_data_from_page()
            filtered_data = [item for item in page_data if item['search_popularity'] >= self.min_popularity_threshold]
            if filtered_data:
                self._append_collected_data(filtered_data)
            return

        # 检查3：切换到检查最后一页前
        if not self.check_pause_state():
            return

        # 第一页没有低价值数据：检查最后一页
        has_low_value_last = self.check_last_page_for_low_value()

        if has_low_value_last:
            # 检查4：开始提取数据前
            if not self.check_pause_state():
                return

            # 最后一页有低价值数据：从第一页开始提取，直到遇到<150
            self.log_ui("最后一页发现小于150的搜索人气")
            self.collect_data_until_low_value()
            return
        else:
            # 检查5：处理三级类目前置检查
            if not self.check_pause_state():
                return

            # 最后一页也没有低价值数据：处理所有三级类目
            self.log_ui("未检测到小于150的搜索人气，开始处理所有三级类目")

            # 再次点击二级类目确保显示其下的三级类目
            self.click_category(2, level2_idx, level2_cat['name'])
            time.sleep(2)

            # 获取三级类目
            self.current_level = 3
            level3_cats = self.get_categories_by_level(3)

            if level3_cats:
                for level3_idx, level3_cat in enumerate(level3_cats, 1):
                    # 检查6：处理每个三级类目前置检查
                    if not self.check_pause_state():
                        break
                    if self.stop_event.is_set():
                        break

                    self.log_ui(f"\n----- 开始处理第{level3_idx}个三级类目: {level3_cat['name']} -----")

                    self.log_ui(f"第{level3_idx}个三级类目提取前，检测目标iframe...")
                    self.detect_target_iframe()

                    # 点击三级类目
                    if not self.click_category(3, level3_idx, level3_cat['name']):
                        continue

                    # 检查7：提取三级类目数据前
                    if not self.check_pause_state():
                        break

                    # 提取三级类目数据（最多6页，遇<150则停）
                    self.collect_data_across_pages()

                # 恢复到二级类目层级
                self.current_level = 2
            else:
                self.log_ui("该二级类目下没有三级类目，提取当前二级类目数据")
                self.collect_data_across_pages()

    def process_secondary_without_tertiary(self):
        """处理无三级类目的二级类目（从第一页开始，最多6页，遇<150则停）"""
        self.log_ui("该二级类目下没有三级类目，开始提取数据")
        self.collect_data_across_pages()

    def collect_data_until_low_value(self):
        """从第一页开始提取，直到遇到<150的数据（用于有三级类目但最后一页有低价值的情况）"""
        try:
            current_page = 1
            max_pages = self.max_pages
            stop_extraction = False
            total_collected = 0

            while current_page <= max_pages and not self.stop_event.is_set() and not stop_extraction:
                # 检查暂停状态
                self.check_pause_state()
                if self.stop_event.is_set():
                    break

                self.log_ui(f"\n===== 开始处理第 {current_page} 页数据 =====")

                # 收集当前页数据
                page_data, found_low_value = self.extract_data_from_page()

                if not page_data:
                    self.log_ui("当前页未提取到任何数据")
                    break

                # 筛选符合条件的数据
                filtered_data = [item for item in page_data if
                                 item['search_popularity'] >= self.min_popularity_threshold]
                total_collected += len(filtered_data)

                self.log_ui(
                    f"第 {current_page} 页，筛选出 {len(filtered_data)} 条符合条件的数据")

                if filtered_data:
                    self._append_collected_data(filtered_data)

                # 遇到低价值数据则停止
                if found_low_value:
                    self.log_ui(f"在第 {current_page} 页发现小于 {self.min_popularity_threshold} 的搜索人气，停止提取")
                    stop_extraction = True
                    break

                # 点击下一页
                self.log_ui("尝试点击下一页...")
                has_next_page = self.click_next_page()

                if not has_next_page:
                    self.log_ui("已到达最后一页，停止处理")
                    break

                current_page += 1
                time.sleep(2)

            self.log_ui(f"共收集到 {total_collected} 条符合条件的数据")
            return total_collected > 0

        except Exception as e:
            self.log_ui(f"处理数据时出错: {str(e)}")
            return False

    def process_level2_category(self):
        """处理二级类目下的所有三级类目"""
        # 记录当前二级类目
        level2_cat = self.selected_categories[2]
        level2_index = level2_cat['index']

        # 检查是否有三级类目
        if level2_cat['has_children']:
            # 有三级类目：先看第一页是否有小于150的
            has_low_value_first = self.check_first_page_for_low_value()

            if has_low_value_first:
                # 第一页有低价值数据，从第一页开始提取当前二级类目数据
                self.log_ui("第一页发现小于150的搜索人气，从第一页开始提取当前二级类目数据")
                data_collected = self.collect_data_across_pages()
                if not data_collected:
                    self.log_ui("当前二级类目数据为空，处理完成")
                    return
            else:
                # 第一页没有低价值数据，再看最后一页
                has_low_value_last = self.check_last_page_for_low_value()

                if has_low_value_last:
                    # 最后一页有低价值数据，从第一页开始提取当前二级类目数据
                    self.log_ui("最后一页发现小于150的搜索人气，从第一页开始提取当前二级类目数据")
                    data_collected = self.collect_data_across_pages()
                    if not data_collected:
                        self.log_ui("当前二级类目数据为空，处理完成")
                        return
                else:
                    # 没有低价值数据，处理所有三级类目
                    self.log_ui("未检测到小于150的搜索人气，处理所有三级类目")

                    # 再次点击二级类目确保显示其下的三级类目
                    self.click_category(2, level2_index, level2_cat['name'])
                    time.sleep(2)

                    # 获取三级类目
                    self.current_level = 3
                    level3_cats = self.get_categories_by_level(3)

                    if level3_cats:
                        for level3_idx, level3_cat in enumerate(level3_cats, 1):
                            # 检查是否已结束
                            if self.stop_event.is_set():
                                break

                            # 检查暂停状态
                            self.check_pause_state()
                            if self.stop_event.is_set():
                                break

                            self.log_ui(f"\n----- 开始处理第{level3_idx}个三级类目: {level3_cat['name']} -----")

                            # 点击三级类目
                            if not self.click_category(3, level3_idx, level3_cat['name']):
                                continue

                            # 提取三级类目数据
                            data_collected = self.collect_data_across_pages()
                            if not data_collected:
                                self.log_ui("当前三级类目数据为空，处理下一个三级类目")
                                continue

                        # 恢复到二级类目层级
                        self.current_level = 2
                    else:
                        self.log_ui("该二级类目下没有三级类目，提取当前二级类目数据")
                        data_collected = self.collect_data_across_pages()
                        if not data_collected:
                            self.log_ui("当前二级类目数据为空，处理完成")
                            return
        else:
            # 没有三级类目，提取这个二级类目的所有大于等于150的数据
            self.log_ui("该二级类目下没有三级类目，提取所有符合条件的数据（最多6页）")
            data_collected = self.collect_data_across_pages()
            if not data_collected:
                self.log_ui("当前二级类目数据为空，处理完成")
                return

    def check_pause_state(self):
        """检查暂停状态"""
        while True:
            with self.lock:
                if self.stop_event.is_set():
                    return False
                if not self.paused:
                    return True

            # 暂停状态下每秒检查一次
            time.sleep(0.1)

    def get_max_page_number(self):
        """获取最大页码数"""
        try:
            # 查找所有页码按钮
            page_elements = self.browser.find_elements(
                By.CSS_SELECTOR,
                "li.ant-pagination-item:not(.ant-pagination-item-active)"
            )

            # 加上当前激活的页码
            active_page = self.browser.find_element(
                By.CSS_SELECTOR,
                "li.ant-pagination-item.ant-pagination-item-active"
            )
            page_elements.append(active_page)

            # 提取所有页码数字
            page_numbers = []
            for elem in page_elements:
                try:
                    num = int(elem.text.strip())
                    page_numbers.append(num)
                except ValueError:
                    continue

            if page_numbers:
                return max(page_numbers)
            else:
                return None

        except NoSuchElementException:
            return None
        except Exception as e:
            self.log_console(f"获取最大页码失败: {str(e)}")
            return None

    def click_page_number(self, page_num):
        """点击指定页码"""
        try:
            # 查找指定页码的按钮
            page_element = self.browser.find_element(
                By.XPATH,
                f"//li[contains(@class, 'ant-pagination-item') and text()='{page_num}']"
            )

            if page_element:
                page_element.click()
                return True
            else:
                return False

        except NoSuchElementException:
            return False
        except Exception as e:
            self.log_console(f"点击页码{page_num}失败: {str(e)}")
            return False

    def create_root_excel_file(self, index, name):
        """创建根类目Excel文件"""
        try:
            # 处理名称中的非法字符
            safe_name = name.replace('/', '-').replace('\\', '-').replace(':', '-').replace('*', '-').replace('?',
                                                                                                              '-').replace(
                '"', '-').replace('<', '-').replace('>', '-').replace('|', '-')

            # 构建文件名
            filename = f"{index}_{safe_name}.xlsx"

            # 获取当前输出目录
            output_dir = self.output_dir or os.getcwd()
            os.makedirs(output_dir, exist_ok=True)
            excel_path = os.path.join(output_dir, filename)

            # 初始化Excel文件
            if self.init_excel(excel_path):
                self.log_ui(f"已创建根类目汇总表格: {excel_path}")
                self.excel_filepath = excel_path
                return excel_path
            else:
                self.log_ui(f"创建根类目表格失败: {excel_path}")
                return None

        except Exception as e:
            self.log_ui(f"创建根类目表格时出错: {str(e)}")
            return None

    def init_excel(self, excel_filename):
        """初始化Excel文件"""
        if self.stop_event.is_set():
            return False

        # 若文件存在，清空所有内容
        if os.path.exists(excel_filename):
            try:
                wb = openpyxl.load_workbook(excel_filename)

                # 清空"生意参谋"工作表
                if "生意参谋" in wb.sheetnames:
                    ws1 = wb["生意参谋"]
                    ws1.delete_rows(1, ws1.max_row)

                wb.save(excel_filename)
                self.log_console(f"已清空Excel汇总表内容：{excel_filename}")
            except Exception as e:
                self.log_console(f"清空Excel文件失败：{e}")
                return False

        # 创建表头
        try:
            wb = openpyxl.load_workbook(excel_filename) if os.path.exists(excel_filename) else openpyxl.Workbook()

            # 处理"生意参谋"工作表
            if "生意参谋" not in wb.sheetnames:
                ws1 = wb.create_sheet(title="生意参谋")
            else:
                ws1 = wb["生意参谋"]
            ws1["A1"] = "搜索词"
            ws1["B1"] = "搜索人气"

            # 处理"衍生关键词"工作表
            if "衍生关键词" not in wb.sheetnames:
                ws2 = wb.create_sheet(title="衍生关键词")
            else:
                ws2 = wb["衍生关键词"]
            ws2["A1"] = "相关关键词"
            ws2["B1"] = "搜索人数"
            ws2["C1"] = "支付转化率"
            ws2["D1"] = "支付人数"
            ws2["E1"] = "需求供给比"
            ws2["F1"] = "天猫商品点击占比"

            # 如果是新建的工作簿，删除默认的Sheet
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

            wb.save(excel_filename)
            return True
        except Exception as e:
            self.log_console(f"初始化Excel表头失败：{e}")
            return False

    def open_excel_file(self):
        """打开当前根类目Excel文件"""
        if not self.current_excel_root or not os.path.exists(self.current_excel_root):
            self.log_ui("根类目Excel文件不存在，无法打开")
            return

        try:
            self.log_ui(f"打开根类目汇总表：{self.current_excel_root}")
            if os.name == 'nt':
                os.startfile(self.current_excel_root)
            else:
                subprocess.run(['open' if os.name == 'posix' else 'xdg-open', self.current_excel_root])
            time.sleep(2)
        except Exception as e:
            self.log_ui(f"打开Excel文件时出错: {str(e)}")

    def connect_to_excel(self):
        """连接到已打开的根类目Excel实例"""
        if not self.current_excel_root:
            return None

        try:
            # 首先检查是否已经有excel_app实例
            if self.excel_app:
                # 验证实例是否仍然有效
                try:
                    _ = self.excel_app.Workbooks.Count
                except:
                    # 实例无效，清除并重新获取
                    self.excel_app = None
            
            # 如果没有有效的excel_app，尝试获取活动实例
            if not self.excel_app:
                try:
                    self.excel_app = win32com.client.GetActiveObject("Excel.Application")
                    self.log_console("已连接到现有Excel实例")
                except Exception as e:
                    self.log_console(f"未找到Excel实例，创建新实例: {str(e)}")
                    self.excel_app = win32com.client.Dispatch("Excel.Application")
                    
                self.excel_app.Visible = True

            # 查找当前根类目Excel工作簿
            workbook = None
            for wb in self.excel_app.Workbooks:
                if os.path.abspath(wb.FullName) == os.path.abspath(self.current_excel_root):
                    workbook = wb
                    break

            # 如果工作簿未打开，则打开它
            if not workbook:
                try:
                    workbook = self.excel_app.Workbooks.Open(self.current_excel_root)
                    time.sleep(2)
                except Exception as e:
                    self.log_ui(f"无法打开Excel文件: {str(e)}")
                    return None

            return workbook
            
        except Exception as e:
            self.log_console(f"连接Excel时出错: {str(e)}")
            return None

    def save_page_data_to_excel(self, page_data, commit=False):
        """将数据写入Excel；commit=True时执行保存"""
        if not self.current_excel_root:
            return False

        has_rows = bool(page_data)
        if not has_rows and not commit:
            return False

        try:
            workbook = self.connect_to_excel()
            if not workbook:
                return self.save_page_data_to_excel_fallback(page_data, commit)

            worksheet = None
            for ws in workbook.Worksheets:
                if ws.Name == "生意参谋":
                    worksheet = ws
                    break
            if not worksheet:
                worksheet = workbook.Worksheets.Add()
                worksheet.Name = "生意参谋"
                worksheet.Cells(1, 1).Value = "搜索词"
                worksheet.Cells(1, 2).Value = "搜索人气"

            if has_rows:
                last_row = worksheet.Cells(worksheet.Rows.Count, "A").End(-4162).Row
                start_row = last_row + 1
                for i, item in enumerate(page_data):
                    row = start_row + i
                    worksheet.Cells(row, 1).Value = item['keyword']
                    worksheet.Cells(row, 2).Value = item['popularity_text']
                self.log_ui(f"已写入{len(page_data)}条数据到Excel（待保存）")
                state = self._get_active_state()
                state.excel_dirty = True

            if commit and self.excel_app:
                try:
                    workbook.Save()
                    state = self._get_active_state()
                    state.excel_dirty = False
                    self.log_ui("已保存Excel文件")
                except Exception as exc:
                    self.log_ui(f"保存Excel失败: {exc}")
                    new_path = os.path.splitext(self.current_excel_root)[0] + "_备份.xlsx"
                    workbook.SaveAs(new_path)
                    self.log_ui(f"文件已另存为: {new_path}")
                    state = self._get_active_state()
                    state.excel_dirty = False
            return True

        except Exception as e:
            self.log_ui(f"写入Excel时出错: {str(e)}，尝试备用方案")
            return self.save_page_data_to_excel_fallback(page_data, commit)

    def save_page_data_to_excel_fallback(self, page_data, commit=False):
        """备用的Excel保存方法（使用openpyxl，自动保存到文件）"""
        if not self.current_excel_root:
            return False
        try:
            wb = openpyxl.load_workbook(self.current_excel_root) if os.path.exists(self.current_excel_root) else openpyxl.Workbook()
            if "生意参谋" not in wb.sheetnames:
                ws = wb.create_sheet(title="生意参谋")
                ws["A1"] = "搜索词"
                ws["B1"] = "搜索人气"
            else:
                ws = wb["生意参谋"]

            if page_data:
                last_row = ws.max_row
                start_row = last_row + 1
                for i, item in enumerate(page_data):
                    row = start_row + i
                    ws.cell(row=row, column=1, value=item['keyword'])
                    ws.cell(row=row, column=2, value=item['popularity_text'])
                self.log_ui(f"备用方法写入{len(page_data)}条数据")

            wb.save(self.current_excel_root)
            state = self._get_active_state()
            state.excel_dirty = False
            if commit:
                self.log_ui("备用方法已保存Excel文件")
            return True
        except Exception as e:
            self.log_ui(f"备用方法保存失败: {str(e)}")
            return False

    def collect_data_across_pages(self):
        """跨分页收集数据（优化停止检查点）"""
        try:
            current_page = 1
            max_pages = self.max_pages
            stop_extraction = False
            total_collected = 0

            while current_page <= max_pages and not self.stop_event.is_set() and not stop_extraction:
                # 【新增】每次处理页面数据前，先检测目标iframe
                self.log_ui(f"\n===== 第{current_page}页数据提取前，检测目标iframe =====")
                has_iframe = self.detect_target_iframe()  # 执行检测

                # 【可选】如果需要在检测到iframe时做特殊处理（如提示用户），可在此添加
                if has_iframe:
                    self.log_ui(f"⚠️ 第{current_page}页提取前检测到目标iframe，请注意验证状态")
                    # 如需暂停提取等待用户处理，可添加：self.check_pause_state()

                # 检查暂停状态（原有逻辑）
                self.check_pause_state()
                if self.stop_event.is_set():
                    break

                self.log_ui(f"===== 开始处理第 {current_page} 页数据 =====")

                # 提取数据前再次检查
                if self.stop_event.is_set():
                    self.log_ui("检测到结束指令，不提取当前页数据")
                    return False

                # 收集当前页数据
                page_data, found_low_value = self.extract_data_from_page()

                # 处理数据前检查
                if self.stop_event.is_set():
                    self.log_ui("检测到结束指令，不处理当前页数据")
                    return False

                # 后续数据处理逻辑（保持不变）...
                # 省略筛选和保存数据的代码...

                current_page += 1
                time.sleep(2)

            self.log_ui(f"共收集到 {total_collected} 条符合条件的数据")
            return True

        except Exception as e:
            self.log_ui(f"处理数据时出错: {str(e)}")
            return False

    def extract_data_from_page(self):
        """从当前页面表格中提取关键词和搜索人气数据（修复转义序列警告）"""
        try:
            try:
                # 检查第一条数据的元素是否存在
                self.browser.find_element(
                    By.CSS_SELECTOR,
                    "tr.ant-table-row.oui-table-row-tree-node-1.ant-table-row-level-0"
                )
            except NoSuchElementException:
                self.log_ui("未找到任何数据行，判定为空数据")
                return [], False  # 返回空数据

            # 等待表格加载完成
            WebDriverWait(self.browser, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "tr.ant-table-row.oui-table-row-tree-node-1"))
            )

            # 使用JavaScript提取数据，修复转义序列警告
            result = self.browser.execute_script("""
                var results = [];
                var foundLowValue = false;
                var minThreshold = arguments[0];
                var debugLogs = []; // 用于调试的日志数组

                // 转换包含"万"的数值为实际数字
                function convertWanValue(valueText) {
                    // 移除所有空格 - 使用双重转义避免Python解释器警告
                    valueText = valueText.replace(/\\s+/g, '');

                    // 检查是否包含"万"
                    if (valueText.includes('万')) {
                        // 提取数字部分
                        var numPart = valueText.replace('万', '');
                        // 尝试转换为浮点数
                        var num = parseFloat(numPart);
                        // 有效数字则乘以10000
                        if (!isNaN(num)) {
                            return num * 10000;
                        }
                    } else {
                        // 普通数字，直接转换
                        var num = parseFloat(valueText);
                        if (!isNaN(num)) {
                            return num;
                        }
                    }
                    return null; // 无法转换
                }

                // 遍历所有符合条件的行
                for (var i = 1; i <= 50; i++) {
                    var rowSelector = 'tr.ant-table-row.oui-table-row-tree-node-' + i + '.ant-table-row-level-0';
                    var row = document.querySelector(rowSelector);

                    if (!row) {
                        continue;
                    }

                    try {
                        // 提取关键词和搜索人气
                        var tds = row.querySelectorAll('td');
                        if (tds.length < 3) {
                            continue;
                        }

                        // 关键词在第二个td
                        var keyword = tds[1].textContent.trim();

                        // 搜索人气在第三个td
                        var popularityElement = tds[2].querySelector('.alife-dt-card-common-table-sortable-value span');
                        if (!popularityElement) {
                            debugLogs.push(`未找到搜索人气元素: 关键词=${keyword}`);
                            continue;
                        }

                        var popularityText = popularityElement.textContent.trim();
                        var popularityValue = 0;
                        var parsingSuccess = true;

                        // 处理范围值，如"1万 ~ 2万"，取最小值
                        if (popularityText.includes('~')) {
                            var parts = popularityText.split('~').map(p => p.trim());

                            // 转换两边的值
                            var value1 = convertWanValue(parts[0]);
                            var value2 = convertWanValue(parts[1]);

                            // 确保两边都是有效数字
                            if (value1 !== null && value2 !== null) {
                                popularityValue = Math.min(value1, value2);
                            } else {
                                parsingSuccess = false;
                                debugLogs.push(`范围值解析失败: "${popularityText}" 部分值无法转换`);
                            }
                        } 
                        // 处理单个带"万"的值，如"1.5万"
                        else if (popularityText.includes('万')) {
                            var value = convertWanValue(popularityText);
                            if (value !== null) {
                                popularityValue = value;
                            } else {
                                parsingSuccess = false;
                                debugLogs.push(`单值解析失败: "${popularityText}" 无法转换`);
                            }
                        }
                        // 处理普通数字值
                        else {
                            var num = parseFloat(popularityText);
                            if (!isNaN(num)) {
                                popularityValue = num;
                            } else {
                                parsingSuccess = false;
                                debugLogs.push(`数字解析失败: "${popularityText}" 不是有效数字`);
                            }
                        }

                        // 调试日志：记录原始文本和解析结果
                        debugLogs.push(`关键词: ${keyword}, 原始值: "${popularityText}", 解析值: ${popularityValue}, 解析成功: ${parsingSuccess}`);

                        // 只有解析成功的情况下才判断是否为低价值
                        if (parsingSuccess) {
                            // 严格判断：只有确实小于阈值才标记
                            if (popularityValue < minThreshold) {
                                foundLowValue = true;
                                debugLogs.push(`>>> 发现低价值数据: ${keyword} (${popularityValue} < ${minThreshold})`);
                            }

                            results.push({
                                keyword: keyword,
                                search_popularity: popularityValue,
                                popularity_text: popularityText
                            });
                        }
                    } catch (e) {
                        debugLogs.push(`提取第${i}行数据失败: ${e.message}`);
                    }
                }

                // 返回结果和调试信息
                return {
                    data: results,
                    foundLowValue: foundLowValue,
                    debugLogs: debugLogs
                };
            """, self.min_popularity_threshold)

            # 输出调试日志到控制台，方便排查问题
            for log in result['debugLogs']:
                self.log_console(log)

            # 验证低价值判断
            if result['foundLowValue']:
                low_count = len([d for d in result['data'] if d['search_popularity'] < self.min_popularity_threshold])
                self.log_ui(f"检测到{low_count}条低价值数据")

            return result['data'], result['foundLowValue']

        except Exception as e:
            self.log_ui(f"提取页面数据时出错: {str(e)}")
            return [], False

    def collect_data_across_pages(self):
        """跨分页收集数据（修复版：增强低价值判断准确性）"""
        try:
            current_page = 1
            max_pages = self.max_pages
            stop_extraction = False
            total_collected = 0  # 记录总收集数量

            while current_page <= max_pages and not self.stop_event.is_set() and not stop_extraction:
                # 检查暂停状态
                self.check_pause_state()
                if self.stop_event.is_set():
                    break

                self.log_ui(f"\n===== 开始处理第 {current_page} 页数据 =====")

                # 收集当前页数据
                page_data, found_low_value = self.extract_data_from_page()

                # 检查是否为空数据
                if not page_data:
                    self.log_ui("当前页未提取到任何数据")
                    break

                # 筛选并保存符合条件的数据（≥150）
                filtered_data = [item for item in page_data if
                                 item['search_popularity'] >= self.min_popularity_threshold]
                total_collected += len(filtered_data)

                self.log_ui(
                    f"第 {current_page} 页共找到 {len(page_data)} 条数据，筛选出 {len(filtered_data)} 条符合条件的数据")

                # 验证低价值判断是否准确
                actual_low_values = [item for item in page_data if
                                     item['search_popularity'] < self.min_popularity_threshold]
                if found_low_value and len(actual_low_values) == 0:
                    self.log_ui(
                        f"警告：检测到低价值标记，但实际未发现小于{self.min_popularity_threshold}的数据，将继续提取")
                    found_low_value = False  # 修正误判

                if filtered_data:
                    self._append_collected_data(filtered_data)

                # 遇到小于阈值的数据则停止提取
                if found_low_value:
                    self.log_ui(
                        f"在第 {current_page} 页发现{len(actual_low_values)}条小于 {self.min_popularity_threshold} 的搜索人气，停止提取当前类目")
                    stop_extraction = True
                    break

                # 达到最大页数限制
                if current_page >= max_pages:
                    self.log_ui(f"已达到最大页数限制 ({max_pages}页)，停止处理")
                    break

                # 点击下一页
                self.log_ui("尝试点击下一页...")
                has_next_page = self.click_next_page()

                if not has_next_page:
                    self.log_ui("已到达最后一页，停止处理")
                    break

                current_page += 1
                time.sleep(2)  # 等待页面加载

            # 检查是否收集到数据
            if total_collected == 0:
                self.log_ui("未收集到任何符合条件的数据")
                return False

            self.log_ui(f"共收集到 {total_collected} 条符合条件的数据")
            return True

        except Exception as e:
            self.log_ui(f"处理数据时出错: {str(e)}")
            return False

    def get_current_page_number(self):
        """获取当前页码"""
        try:
            # 查找当前激活的页码元素
            page_element = self.browser.find_element(
                By.CSS_SELECTOR,
                "li.ant-pagination-item.ant-pagination-item-active"
            )
            return int(page_element.text)
        except Exception as e:
            self.log_console(f"获取当前页码失败: {str(e)}")
            return 1  # 默认返回1

    def click_next_page(self):
        """点击下一页（新增备用方案：点击页码翻页）"""
        try:
            # 主要方案：尝试点击"下一页"按钮
            WebDriverWait(self.browser, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "li.ant-pagination-next"))
            )

            # 使用完整的CSS选择器点击下一页
            selector = "li.ant-pagination-next:not(.ant-pagination-disabled)"
            next_page = self.browser.find_element(By.CSS_SELECTOR, selector)

            # 检查是否禁用
            if "ant-pagination-disabled" in next_page.get_attribute("class"):
                self.log_ui("下一页按钮已禁用，尝试备用方案")
                return self._next_page_fallback()  # 调用备用方案

            # 点击下一页
            try:
                next_page.click()
                self.log_ui("使用主要方案点击下一页成功")
                time.sleep(1)
                return True
            except:
                # 尝试点击内部的a标签
                link = next_page.find_element(By.TAG_NAME, "a")
                link.click()
                self.log_ui("使用主要方案点击下一页成功")
                time.sleep(1)
                return True

        except NoSuchElementException:
            self.log_ui("未找到下一页按钮，尝试备用方案（点击页码）")
            return self._next_page_fallback()
        except Exception as e:
            self.log_ui(f"点击下一页时出错: {str(e)}，尝试备用方案（点击页码）")
            return self._next_page_fallback()

    def _next_page_fallback(self):
        """备用翻页方案：通过点击页码翻页"""
        try:
            # 获取当前页码
            current_page = self.get_current_page_number()
            next_page_num = current_page + 1

            # 检查是否超过最大页数限制
            if next_page_num > self.max_pages:
                self.log_ui(f"已达到最大页数限制 ({self.max_pages}页)，无法继续翻页")
                return False

            self.log_ui(f"尝试点击页码 {next_page_num} 进行翻页")
            success = self.click_page_number(next_page_num)

            if success:
                time.sleep(2)  # 等待页面加载
                self.log_ui(f"备用方案：成功切换到第 {next_page_num} 页")
                return True
            else:
                self.log_ui(f"备用方案：点击页码 {next_page_num} 失败")
                return False
        except Exception as e:
            self.log_ui(f"备用翻页方案出错: {str(e)}")
            return False


    def cookie_login(self, session_id):
        """使用保存的Cookie登录指定会话"""
        state = self.session_states.get(session_id)
        if not state or not state.browser:
            self.log_ui(f"窗口{session_id + 1} 请先打开浏览器")
            return

        cookie_path = self.cookie_path
        if not os.path.exists(cookie_path):
            alt_path = os.path.join(os.getcwd(), "cookie.txt")
            if os.path.exists(alt_path):
                cookie_path = alt_path
            else:
                self.log_ui(f"未找到cookie.txt（尝试路径: {self.cookie_path}）")
                return

        with open(cookie_path, 'r', encoding='utf-8') as f:
            raw = f.read().strip().split('; ')

        cookies = []
        for item in raw:
            if '=' in item:
                name, value = item.split('=', 1)
                cookies.append((name, value))

        if not cookies:
            self.log_ui("Cookie文件无有效内容，无法登录")
            return

        self.log_ui(f"窗口{session_id + 1} 开始写入 {len(cookies)} 条Cookie")
        self._run_in_session_thread(state, self._apply_cookies_to_session, cookies)

    def _apply_cookies_to_session(self, cookies):
        state = self._get_active_state()
        try:
            self.browser.get("https://sycm.taobao.com/")
            for name, value in cookies:
                try:
                    self.browser.add_cookie({
                        'name': name,
                        'value': value,
                        'domain': '.taobao.com',
                        'path': '/',
                        'httpOnly': False,
                        'secure': False
                    })
                except Exception as exc:  # pylint: disable=broad-except
                    self.log_console(f"[窗口{state.session_id + 1}] 写入Cookie失败 ({name}): {exc}")

            self.browser.refresh()
            self.log_session("Cookie登录完成，可继续操作")
        except Exception as exc:  # pylint: disable=broad-except
            self.log_session(f"Cookie登录失败: {exc}")

    def save_cookie(self):
        """保存当前登录状态的Cookie"""
        target_state = next((state for state in self._iter_real_sessions() if state.browser), None)
        if not target_state:
            self.log_ui("请先打开并登录至少一个浏览器窗口")
            return

        def do_save():
            try:
                state = self._get_active_state()
                self.log_session("正在保存Cookie...")
                cookies = self.browser.get_cookies()
                self.log_console(f"[窗口{state.session_id + 1}] 获取到{len(cookies)}条Cookie")

                cookie_str = '; '.join([f"{cookie['name']}={cookie['value']}" for cookie in cookies])

                with open(self.cookie_path, 'w', encoding='utf-8') as f:
                    f.write(cookie_str)

                self.log_session(f"Cookie已保存到 {self.cookie_path}，共{len(cookies)}条")
            except Exception as exc:  # pylint: disable-broad-except
                self.log_session(f"保存Cookie失败: {exc}")

        self._run_in_session_thread(target_state, do_save)

    def toggle_pause(self, session_id):
        """切换指定会话的暂停/继续状态"""
        state = self.session_states.get(session_id)
        if not state or not state.processing:
            self.log_ui(f"窗口{session_id + 1} 当前没有运行中的任务")
            return

        panel = getattr(self, "session_panels", {}).get(session_id)
        if state.paused:
            state.pause_event.set()
            state.paused = False
            if panel:
                panel["pause_btn"].config(text="暂停")
            self.set_session_status(session_id, "运行中")
            self.log_ui(f"窗口{session_id + 1} 已恢复运行")
        else:
            state.pause_event.clear()
            state.paused = True
            if panel:
                panel["pause_btn"].config(text="继续")
            self.set_session_status(session_id, "已暂停")
            self.log_ui(f"窗口{session_id + 1} 已暂停")

    def stop_processing(self, session_id):
        """停止指定会话的处理"""
        state = self.session_states.get(session_id)
        if not state or not state.processing:
            self.log_ui(f"窗口{session_id + 1} 当前没有进行中的任务")
            return

        state.stop_event.set()
        state.pause_event.set()
        state.paused = False
        state.processing = False

        panel = getattr(self, "session_panels", {}).get(session_id)
        if panel:
            panel["pause_btn"].config(state=tk.DISABLED, text="暂停")
            panel["stop_btn"].config(state=tk.DISABLED)
        self.set_session_status(session_id, "准备完成")
        self.log_ui(f"窗口{session_id + 1} 已发出停止指令，将尽快结束当前任务")


    def release_excel_connection(self):
        """仅释放Excel/WPS连接，不关闭文件，允许用户手动操作"""
        self.log_ui("释放Excel/WPS程序连接...")

        if self.excel_app:
            try:
                # 仅释放COM对象引用，不关闭工作簿和程序
                self.excel_app = None
                # 强制Python释放COM对象
                import gc
                gc.collect()
                self.log_ui("已释放Excel/WPS连接，文件保持打开状态")
            except Exception as e:
                self.log_ui(f"释放连接时出错: {str(e)}")

    def copy_matched_excel_files(self):
        """将Excel文件分配到对应目标文件夹（适配UI日志输出）"""
        try:
            # 源目录：当前文件夹
            source_dir = os.getcwd()
            # 目标根目录（可根据需要修改）
            target_root = r"D:\生意参谋\搜索词"

            # 确保目标根目录存在
            if not os.path.exists(target_root):
                self.log_ui(f"错误：目标根目录不存在 - {target_root}")
                btn = getattr(self, "distribute_btn", None)
                if btn:
                    btn.config(state=tk.NORMAL)
                return

            # 正则表达式：匹配Excel文件名（提取序号）
            excel_pattern = re.compile(r'^(\d+)_.*?\.(xlsx|xls|xlsm|xlsb)$', re.IGNORECASE)
            # 正则表达式：匹配目标文件夹名（提取序号列表）
            folder_pattern = re.compile(r'=(\d+(?:,\d+)*)$')

            # 收集所有Excel文件的序号与路径映射
            excel_map = {}
            for filename in os.listdir(source_dir):
                source_path = os.path.join(source_dir, filename)
                if os.path.isfile(source_path):
                    match = excel_pattern.match(filename)
                    if match:
                        serial = match.group(1)  # 提取文件序号
                        if serial in excel_map:
                            excel_map[serial].append({"path": source_path, "name": filename})
                        else:
                            excel_map[serial] = [{"path": source_path, "name": filename}]

            self.log_ui(f"已识别 {len(excel_map)} 组带序号的Excel文件")
            total_copied = 0

            # 遍历目标根目录下的所有文件夹
            for folder_name in os.listdir(target_root):
                folder_path = os.path.join(target_root, folder_name)
                if not os.path.isdir(folder_path):
                    continue

                # 从文件夹名称中提取序号列表
                folder_match = folder_pattern.search(folder_name)
                if not folder_match:
                    continue

                # 分割序号列表
                serials = folder_match.group(1).split(',')
                self.log_ui(f"处理文件夹 [{folder_name}]，包含序号: {','.join(serials)}")

                # 复制匹配的文件
                for serial in serials:
                    if serial in excel_map:
                        for excel_info in excel_map[serial]:
                            source_file = excel_info["path"]
                            target_file = os.path.join(folder_path, excel_info["name"])

                            try:
                                # 复制文件（保留元数据）
                                shutil.copy2(source_file, target_file)
                                self.log_ui(f"已复制: {excel_info['name']} -> {folder_name}")
                                total_copied += 1
                            except Exception as e:
                                self.log_ui(f"复制失败 [{excel_info['name']}] 到 [{folder_name}]: {str(e)}")

            self.log_ui(f"\n文件分配完成，共成功复制 {total_copied} 个文件")

        except Exception as e:
            self.log_ui(f"文件分配过程出错: {str(e)}")
            traceback.print_exc()
        finally:
            # 恢复按钮状态
            btn = getattr(self, "distribute_btn", None)
            if btn:
                btn.config(state=tk.NORMAL)

    # def release_office_resources(self):
    #     """优雅释放Excel/WPS资源，避免强制占用"""
    #     self.log_ui("开始释放Office/WPS资源...")
    #
    #     # 1. 释放Excel实例（如果存在）
    #     if self.excel_app:
    #         try:
    #             # 先尝试正常关闭所有工作簿
    #             for wb in self.excel_app.Workbooks:
    #                 try:
    #                     wb.Close(SaveChanges=False)  # 不保存强制关闭（避免弹窗阻塞）
    #                 except:
    #                     pass
    #             # 退出应用
    #             self.excel_app.Quit()
    #             self.log_ui("Excel/WPS实例已正常退出")
    #         except Exception as e:
    #             self.log_ui(f"正常关闭失败，尝试释放COM对象: {str(e)}")
    #             # 强制释放COM对象（解决顽固占用）
    #             import pythoncom
    #             pythoncom.CoUninitialize()
    #         finally:
    #             self.excel_app = None  # 清空引用
    #
    #     # 2. 兜底清理WPS进程（仅在正常释放失败后使用）
    #     try:
    #         # 先检查是否有残留进程
    #         import psutil
    #         wps_processes = ["wps.exe", "et.exe", "wpscenter.exe"]
    #         for proc_name in wps_processes:
    #             for proc in psutil.process_iter(['name']):
    #                 if proc.info['name'].lower() == proc_name.lower():
    #                     # 先尝试优雅终止
    #                     proc.terminate()
    #                     # 等待1秒，若未终止则强制杀死
    #                     if proc.is_alive():
    #                         time.sleep(1)
    #                         if proc.is_alive():
    #                             proc.kill()
    #                             self.log_ui(f"已强制终止残留进程: {proc_name}")
    #                         else:
    #                             self.log_ui(f"已正常终止进程: {proc_name}")
    #                     else:
    #                         self.log_ui(f"已正常终止进程: {proc_name}")
    #     except ImportError:
    #         self.log_ui("未安装psutil，使用传统方式清理进程")
    #         # 传统方式（仅作为备选，可能仍有残留）
    #         os.system("taskkill /f /im wps.exe >nul 2>&1")
    #         os.system("taskkill /f /im et.exe >nul 2>&1")
    #         os.system("taskkill /f /im wpscenter.exe >nul 2>&1")
    #     except Exception as e:
    #         self.log_ui(f"清理进程时出错: {str(e)}")

    def on_closing(self):
        """窗口关闭时的处理，避免阻塞UI"""
        if self.closing:
            return
        self.closing = True
        self.log_ui("正在关闭程序...")

        for state in self._iter_real_sessions():
            state.stop_event.set()
            state.pause_event.set()

        self.stop_event.set()
        self.pause_event.set()

        self.after(0, self.destroy)


    # def kill_wps_processes(self):
    #     """强制关闭残留的WPS进程"""
    #     try:
    #         # 关闭WPS表格进程
    #         os.system("taskkill /f /im wps.exe >nul 2>&1")
    #         os.system("taskkill /f /im et.exe >nul 2>&1")  # WPS表格进程名
    #         os.system("taskkill /f /im wpscenter.exe >nul 2>&1")
    #         self.log_console("已清理WPS相关进程")
    #     except Exception as e:
    #         self.log_console(f"清理WPS进程时出错: {str(e)}")


if __name__ == "__main__":
    try:
        app = CategoryAutoExtractor()
        app.mainloop()
    except Exception as e:
        print(f"程序启动失败: {str(e)}")
        traceback.print_exc()
